#!/usr/bin/env python
"""Extract parameters from a Drawdown excel model to help create
   the Python implementation of the solution and its scenarios.
   This file is meant to be used as a script; `python solution_xls_extract.py --help` for instructions

   The code in this file is licensed under the GNU AFFERO GENERAL PUBLIC LICENSE
   version 3.0.

   Outputs of this utility are considered to be data and do not automatically
   carry the license used for the code in this utility. It is up to the user and
   copyright holder of the inputs to determine what copyright applies to the
   output.
"""

import argparse
import datetime
import hashlib
import json
import os.path
from pathlib import Path
import re
import sys
import unicodedata
import warnings

import openpyxl
import numpy as np
import pandas as pd
import pytest

from tools.util import convert_bool, xls, xli, xln, co, find_in_column
from tools.vma_xls_extract import VMAReader
from model import advanced_controls as ac


pd.set_option('display.max_rows', 500)
pd.set_option('display.max_columns', 500)
pd.set_option('display.width', 1000)
warnings.filterwarnings("ignore",module=".*openpyxl.*")

# Some constants to monitor for updates:

# Zero adoption solutions are solutions that legitimately go to zero.  Used to distinguish
# "actual zero" from zero used as "no data"
zero_adoption_solutions = ['nuclear', 'cars', 'geothermal', 'improvedcookstoves', 'waterefficiency']


# Type of solution
is_rrs = False
is_land = False
is_elecgen = False


# convert_sr_float parses the combined value + formula content that occurs in the ScenarioRecord tab

def convert_sr_float(tab, row=None, col=None):
    """Return floating point value from Excel ScenarioRecord tab.

       There are three main formats:
       + simple: 0.182810601365724
       + percentage: 20%
       + annotated: Val:(0.182810601365724) Formula:='Variable Meta-analysis'!G1411

       Accepts either (tab, row, col) arguments or a single (value) argument
    """
    if row is not None:
        val = xls(tab, row, col)
    else:
        val = tab
    
    m = re.match(r'Val:\(([-+]?(\d+(\.\d*)?|\d+(\,\d*)?|\.\d+)([eE][-+]?\d+)?)\) Formula:=',
                 str(val))
    if m:
        s = str(m.group(1)).replace(',', '.')
        return float(s)
    if str(val).startswith('Val:() Formula:='):
        return float(0.0)
    if str(val).endswith('%'):
        (num, _) = str(val).split('%', maxsplit=1)
        return float(num) / 100.0
    if val == '':
        return 0.0
    return float(val)


def get_rrs_scenarios(wb, solution_category):
    """Extract scenarios from an RRS Excel file.
       Arguments:
         wb: Excel workbook as returned by openpyxl.
         solution_category: to populate in the scenario
       Returns:
         dict of scenario name to advanced controls dict, suitable for writing to ac/*.json files.
    """
    sr_tab = wb['ScenarioRecord']
    scenarios = {}
    for row in range(sr_tab.min_row, sr_tab.max_row):
        col_d = xls(sr_tab, row, co("D"))
        col_e = xls(sr_tab, row, co("E"))
        if col_d == 'Name of Scenario:' and 'TEMPLATE' not in col_e:
            # start of scenario block
            scenario_name = col_e
            if 'broken' in scenario_name:
                continue
            s = {}

            s['name'] = scenario_name
            s['solution_category'] = solution_category
            s['vmas'] = 'VMAs'

            # Note this is hidden text.
            s['creation_date'] = xls(sr_tab, row, co("B"))  
            # Throw an exception if the date is not in the expected format.
            _scenario_creation_date_from_str(s['creation_date'])
            s['description'] = xls(sr_tab, row + 1, co("E"))

            report_years = xls(sr_tab, row + 2, co("E"))  # E:2 from top of scenario
            (start, end) = report_years.split('-')
            s['report_start_year'] = int(start)
            s['report_end_year'] = int(end)

            assert xls(sr_tab, row + 46, co("B")) == 'Conventional'
            assert xls(sr_tab, row + 47, co("D")) == 'First Cost:'
            s['conv_2014_cost'] = link_vma(sr_tab, row + 47, co("E"))
            s['conv_first_cost_efficiency_rate'] = convert_sr_float(sr_tab, row + 48, co("E"))
            s['conv_lifetime_capacity'] = link_vma(sr_tab, row + 49, co("E"))
            s['conv_avg_annual_use'] = link_vma(sr_tab, row + 50, co("E"))
            s['conv_var_oper_cost_per_funit'] = link_vma(sr_tab, row + 51, co("E"))
            s['conv_fixed_oper_cost_per_iunit'] = link_vma(sr_tab, row + 52, co("E"))
            s['conv_fuel_cost_per_funit'] = convert_sr_float(sr_tab, row + 54, co("E"))

            assert xls(sr_tab, row + 64, co("B")) == 'Solution'
            assert xls(sr_tab, row + 65, co("D")) == 'First Cost:'
            s['pds_2014_cost'] = s['ref_2014_cost'] = link_vma(sr_tab, row + 65, co("E"))
            s['soln_first_cost_efficiency_rate'] = convert_sr_float(sr_tab, row + 66, co("E"))
            s['soln_first_cost_below_conv'] = convert_bool(xls(sr_tab, row + 66, co("G")))
            s['soln_lifetime_capacity'] = link_vma(sr_tab, row + 67, co("E"))
            s['soln_avg_annual_use'] = link_vma(sr_tab, row + 68, co("E"))
            s['soln_var_oper_cost_per_funit'] = link_vma(sr_tab, row + 69, co("E"))
            s['soln_fixed_oper_cost_per_iunit'] = link_vma(sr_tab, row + 70, co("E"))
            s['soln_fuel_cost_per_funit'] = convert_sr_float(sr_tab, row + 72, co("E"))

            assert xls(sr_tab, row + 76, co("B")) == 'General'
            s['npv_discount_rate'] = convert_sr_float(sr_tab, row + 77, co("E"))

            assert xls(sr_tab, row + 86, co("B")) == 'EMISSIONS INPUTS'

            assert xls(sr_tab, row + 88, co("B")) == 'Grid Emissions'
            s['conv_annual_energy_used'] = link_vma(sr_tab, row + 89, co("E"))
            s['soln_energy_efficiency_factor'] = link_vma(sr_tab, row + 90, co("E"))
            s['soln_annual_energy_used'] = link_vma(sr_tab, row + 91, co("E"))

            assert xls(sr_tab, row + 94, co("B")) == 'Fuel Emissions'
            s['conv_fuel_consumed_per_funit'] = link_vma(sr_tab, row + 95, co("E"))
            s['soln_fuel_efficiency_factor'] = link_vma(sr_tab, row + 96, co("E"))
            s['conv_fuel_emissions_factor'] = link_vma(sr_tab, row + 97, co("E"))
            s['soln_fuel_emissions_factor'] = link_vma(sr_tab, row + 98, co("E"))

            assert xls(sr_tab, row + 103, co("B")) == 'Direct Emissions'
            s['conv_emissions_per_funit'] = link_vma(sr_tab, row + 105, co("E"))
            s['soln_emissions_per_funit'] = link_vma(sr_tab, row + 106, co("E"))

            assert xls(sr_tab, row + 111, co("B")) == 'Indirect Emissions'
            s['conv_indirect_co2_per_unit'] = link_vma(sr_tab, row + 112, co("E"))
            s['soln_indirect_co2_per_iunit'] = link_vma(sr_tab, row + 113, co("E"))
            i_vs_f = xls(sr_tab, row + 114, co("E")).lower()
            s['conv_indirect_co2_is_iunits'] = False if i_vs_f == 'functional' else True

            assert xls(sr_tab, row + 118, co("B")) == 'Optional Inputs'
            s['ch4_co2_per_funit'] = link_vma(sr_tab, row + 119, co("E"))
            s['ch4_is_co2eq'] = ("CO2eq" in xls(sr_tab, row + 119, co("F")))
            s['n2o_co2_per_funit'] = link_vma(sr_tab, row + 120, co("E"))
            s['n2o_is_co2eq'] = ("CO2eq" in xls(sr_tab, row + 120, co("F")))
            s['co2eq_conversion_source'] = xls(sr_tab, row + 121, co("E"))

            assert xls(sr_tab, row + 124, co("B")) == 'General Climate Inputs'
            s['emissions_use_co2eq'] = convert_bool(xls(sr_tab, row + 125, co("E")))
            s['emissions_grid_source'] = xls(sr_tab, row + 126, co("E"))
            s['emissions_grid_range'] = xls(sr_tab, row + 127, co("E"))

            assert xls(sr_tab, row + 135, co("B")) == 'TAM'
            s['source_until_2014'] = normalize_source_name(xls(sr_tab, row + 136, co("E")))
            s['ref_source_post_2014'] = normalize_source_name(xls(sr_tab, row + 136, co("H")))
            s['pds_source_post_2014'] = normalize_source_name(xls(sr_tab, row + 136, co("K")))

            s['ref_base_adoption'] = {
                'World': convert_sr_float(sr_tab, row + 151, co("E")),
                'OECD90': convert_sr_float(sr_tab, row + 152, co("E")),
                'Eastern Europe': convert_sr_float(sr_tab, row + 153, co("E")),
                'Asia (Sans Japan)': convert_sr_float(sr_tab, row + 154, co("E")),
                'Middle East and Africa': convert_sr_float(sr_tab, row + 155, co("E")),
                'Latin America': convert_sr_float(sr_tab, row + 156, co("E")),
                'China': convert_sr_float(sr_tab, row + 157, co("E")),
                'India': convert_sr_float(sr_tab, row + 158, co("E")),
                'EU': convert_sr_float(sr_tab, row + 159, co("E")),
                'USA': convert_sr_float(sr_tab, row + 160, co("E")),
            }

            assert xls(sr_tab, row + 163, co("B")) == 'PDS ADOPTION SCENARIO INPUTS'
            s['soln_pds_adoption_basis'] = xls(sr_tab, row + 164, co("E"))
            s['soln_pds_adoption_regional_data'] = convert_bool(xls(sr_tab, row + 165, co("E")))

            s['pds_adoption_final_percentage'] = {
                'World': xln(sr_tab, row + 170, co("E")),
                'OECD90': xln(sr_tab, row + 171, co("E")),
                'Eastern Europe': xln(sr_tab, row + 172, co("E")),
                'Asia (Sans Japan)': xln(sr_tab, row + 173, co("E")),
                'Middle East and Africa': xln(sr_tab, row + 174, co("E")),
                'Latin America': xln(sr_tab, row + 175, co("E")),
                'China': xln(sr_tab, row + 176, co("E")),
                'India': xln(sr_tab, row + 177, co("E")),
                'EU': xln(sr_tab, row + 178, co("E")),
                'USA': xln(sr_tab, row + 179, co("E")),
            }

            if s['soln_pds_adoption_basis'] == 'DEFAULT S-Curve':
                s_curve_type = xls(sr_tab, row + 181, co("E"))
                if s_curve_type == 'Alternate S-Curve (Bass Model)':
                    s['soln_pds_adoption_basis'] = 'Bass Diffusion S-Curve'
                    s['pds_adoption_s_curve_innovation'] = {
                        'World': convert_sr_float(sr_tab, row + 170, co("G")),
                        'OECD90': convert_sr_float(sr_tab, row + 171, co("G")),
                        'Eastern Europe': convert_sr_float(sr_tab, row + 172, co("G")),
                        'Asia (Sans Japan)': convert_sr_float(sr_tab, row + 173, co("G")),
                        'Middle East and Africa': convert_sr_float(sr_tab, row + 174, co("G")),
                        'Latin America': convert_sr_float(sr_tab, row + 175, co("G")),
                        'China': convert_sr_float(sr_tab, row + 176, co("G")),
                        'India': convert_sr_float(sr_tab, row + 177, co("G")),
                        'EU': convert_sr_float(sr_tab, row + 178, co("G")),
                        'USA': convert_sr_float(sr_tab, row + 179, co("G"))}
                    s['pds_adoption_s_curve_imitation'] = {
                        'World': convert_sr_float(sr_tab, row + 170, co("H")),
                        'OECD90': convert_sr_float(sr_tab, row + 171, co("H")),
                        'Eastern Europe': convert_sr_float(sr_tab, row + 172, co("H")),
                        'Asia (Sans Japan)': convert_sr_float(sr_tab, row + 173, co("H")),
                        'Middle East and Africa': convert_sr_float(sr_tab, row + 174, co("H")),
                        'Latin America': convert_sr_float(sr_tab, row + 175, co("H")),
                        'China': convert_sr_float(sr_tab, row + 176, co("H")),
                        'India': convert_sr_float(sr_tab, row + 177, co("H")),
                        'EU': convert_sr_float(sr_tab, row + 178, co("H")),
                        'USA': convert_sr_float(sr_tab, row + 179, co("H"))}
                elif s_curve_type == 'Default S-Curve (Logistic Model)':
                    s['soln_pds_adoption_basis'] = 'Logistic S-Curve'
                else:
                    raise ValueError('Unknown S-Curve:' + s_curve_type)

            assert xls(sr_tab, row + 183, co("B")) == 'Existing PDS Prognostication Assumptions'
            adopt = normalize_source_name(xls(sr_tab, row + 184, co("E")))
            adopt = normalize_case_name(adopt)
            if adopt: 
                s['soln_pds_adoption_prognostication_source'] = adopt
            
            adopt = xls(sr_tab, row + 185, co("E"))
            if adopt: 
                s['soln_pds_adoption_prognostication_trend'] = adopt
            
            adopt = xls(sr_tab, row + 186, co("E"))
            if adopt: 
                s['soln_pds_adoption_prognostication_growth'] = adopt

            assert xls(sr_tab, row + 194, co("B")) == 'Fully Customized PDS'
            custom = xls(sr_tab, row + 195, co("E"))
            if custom:
                s['soln_pds_adoption_custom_name'] = custom
                if 'soln_pds_adoption_basis' not in s:  # sometimes row 164 is blank
                    s['soln_pds_adoption_basis'] = 'Fully Customized PDS'
                try:
                    values = xls(sr_tab, row + 196, co("E")).split(',')
                    s['soln_pds_adoption_scenarios_included'] = [ int(x.strip())-1 for x in values if x.strip() != '' ]
                except ValueError:
                    # TODO: Warn?
                    s['soln_pds_adoption_scenarios_included'] = list(range(10))

            assert xls(sr_tab, row + 198, co("B")) == 'REF ADOPTION SCENARIO INPUTS'
            adopt = xls(sr_tab, row + 199, co("E"))
            if adopt: 
                s['soln_ref_adoption_basis'] = adopt
            
            custom = xls(sr_tab, row + 200, co("E"))
            if custom: 
                s['soln_ref_adoption_custom_name'] = custom
            s['soln_ref_adoption_regional_data'] = convert_bool(xls(sr_tab, row + 201, co("E")))

            assert xls(sr_tab, row + 217, co("B")) == 'Adoption Adjustment'
            adjust = xls(sr_tab, row + 218, co("E"))
            if adjust and adjust != "(none)":
                s['pds_adoption_use_ref_years'] = [int(x) for x in adjust.split(',') if x != '']
            adjust = xls(sr_tab, row + 219, co("E"))
            if adjust and adjust != "(none)":
                s['ref_adoption_use_pds_years'] = [int(x) for x in adjust.split(',') if x != '']

            row += 202
            scenarios[scenario_name] = s
    return scenarios


def get_land_scenarios(wb, solution_category):
    """Extract scenarios from a LAND Excel file.
       Arguments:
         wb: Excel workbook returned by openpyxl.
         solution_category: to populate in the scenario
       Return:
         dict of scenario name to advanced controls dict, suitable for writing to ac/*.json files.
    """
    sr_tab = wb['ScenarioRecord']
    scenarios = {}
    for row in range(sr_tab.min_row, sr_tab.max_row):
        col_d = xls(sr_tab, row, co("D"))
        col_e = xls(sr_tab, row, co("E"))
        if col_d == 'Name of Scenario:' and 'TEMPLATE' not in col_e:
            # start of scenario block
            scenario_name = col_e
            if 'broken' in scenario_name:
                continue
            s = {}

            s['name'] = scenario_name
            s['solution_category'] = solution_category
            s['vmas'] = 'VMAs'

            # Note this is hidden text.
            s['creation_date'] = xls(sr_tab, row, co("B"))  
            # Throw an exception if the date is not in the expected format.
            _scenario_creation_date_from_str(s['creation_date'])
            
            s['description'] = xls(sr_tab, row + 1, co("E"))
            report_years = xls(sr_tab, row + 2, co("E"))
            (start, end) = report_years.split('-')
            s['report_start_year'] = int(start)
            s['report_end_year'] = int(end)

            assert xls(sr_tab, row + 201, co("D")) == 'Custom TLA Used?:'
            s['use_custom_tla'] = convert_bool(xls(sr_tab, row + 201, co("E")))
            has_values = (xls(sr_tab, row + 203, co("E")) and 
                    xls(sr_tab, row + 203, co("H")) and
                    xls(sr_tab, row + 203, co("K")))
            if s['use_custom_tla'] and has_values:
                world_tla = convert_sr_float(xls(sr_tab, row + 203, co("E")))
                future_ref = convert_sr_float(xls(sr_tab, row + 203, co("H")))
                future_pds = convert_sr_float(xls(sr_tab, row + 203, co("K")))
                # we can only handle if all three are the same
                err = f"mismatched Custom TLA values at row {row + 203}"
                assert world_tla == pytest.approx(future_ref), err
                assert world_tla == pytest.approx(future_pds), err
                s['custom_tla_fixed_value'] = world_tla

            s['ref_base_adoption'] = {
                'World': convert_sr_float(sr_tab, row + 218, co("E")),
                'OECD90': convert_sr_float(sr_tab, row + 219, co("E")),
                'Eastern Europe': convert_sr_float(sr_tab, row + 220, co("E")),
                'Asia (Sans Japan)': convert_sr_float(sr_tab, row + 221, co("E")),
                'Middle East and Africa': convert_sr_float(sr_tab, row + 222, co("E")),
                'Latin America': convert_sr_float(sr_tab, row + 223, co("E")),
                'China': convert_sr_float(sr_tab, row + 224, co("E")),
                'India': convert_sr_float(sr_tab, row + 225, co("E")),
                'EU': convert_sr_float(sr_tab, row + 226, co("E")),
                'USA': convert_sr_float(sr_tab, row + 227, co("E")),
            }

            assert xls(sr_tab, row + 230, co("B")) == 'PDS ADOPTION SCENARIO INPUTS'
            adopt = xls(sr_tab, row + 231, co("E"))
            if adopt: 
                s['soln_pds_adoption_basis'] = adopt
            s['soln_pds_adoption_regional_data'] = convert_bool(xls(sr_tab, row + 232, co("E")))

            s['pds_adoption_final_percentage'] = {
                'World': xln(sr_tab, row + 236, co("E")),
                'OECD90': xln(sr_tab, row + 237, co("E")),
                'Eastern Europe': xln(sr_tab, row + 238, co("E")),
                'Asia (Sans Japan)': xln(sr_tab, row + 239, co("E")),
                'Middle East and Africa': xln(sr_tab, row + 240, co("E")),
                'Latin America': xln(sr_tab, row + 241, co("E")),
                'China': xln(sr_tab, row + 242, co("E")),
                'India': xln(sr_tab, row + 243, co("E")),
                'EU': xln(sr_tab, row + 244, co("E")),
                'USA': xln(sr_tab, row + 245, co("E")),
            }

            assert xls(sr_tab, row + 258, co("B")) == 'Fully Customized PDS'
            custom = xls(sr_tab, row + 259, co("E"))
            if custom:
                s['soln_pds_adoption_custom_name'] = custom
                if 'soln_pds_adoption_basis' not in s:  # sometimes row 164 is blank
                    s['soln_pds_adoption_basis'] = 'Fully Customized PDS'
                try:
                    values = xls(sr_tab, row + 260, co("E")).split(',')
                    s['soln_pds_adoption_scenarios_included'] = [ int(x.strip())-1 for x in values if x.strip() != '' ]
                except ValueError:
                    # TODO: Warn?
                    s['soln_pds_adoption_scenarios_included'] = list(range(10))
                
                try:
                    high, low = xls(sr_tab, row + 260, co("H")).split(',')
                except ValueError:
                    high = low = "1"
                if high != "1" and high != "":
                    s['soln_pds_adoption_custom_high_sd_mult'] = float(high)
                if low != "1" and low != "":
                    s['soln_pds_adoption_custom_low_sd_mult'] = float(low)

            assert xls(sr_tab, row + 262, co("B")) == 'REF ADOPTION SCENARIO INPUTS'
            adopt = xls(sr_tab, row + 263, co("E"))
            if adopt: 
                s['soln_ref_adoption_basis'] = adopt
            
            custom = xls(sr_tab, row + 264, co("E"))
            if custom: 
                s['soln_ref_adoption_custom_name'] = custom
            
            s['soln_ref_adoption_regional_data'] = convert_bool(xls(sr_tab, row + 265, co("E")))

            assert xls(sr_tab, row + 286, co("B")) == 'Adoption Adjustment'
            adjust = xls(sr_tab, row + 287, co("E"))
            if adjust and adjust != "(none)":
                s['pds_adoption_use_ref_years'] = [int(x) for x in adjust.split(',') if x != '']
            
            adjust = xls(sr_tab, row + 288, co("E"))
            if adjust and adjust != "(none)":
                s['ref_adoption_use_pds_years'] = [int(x) for x in adjust.split(',') if x != '']
            # TODO: handle soln_pds_adoption_prognostication_source

            assert xls(sr_tab, row + 54, co("B")) == 'Conventional'
            assert xls(sr_tab, row + 55, co("D")) == 'First Cost:'
            s['conv_2014_cost'] = link_vma(sr_tab, row + 55, co("E"))
            s['conv_first_cost_efficiency_rate'] = 0.0  # always 0 for LAND models
            s['conv_fixed_oper_cost_per_iunit'] = link_vma(sr_tab, row + 56, co("E"))
            s['conv_expected_lifetime'] = convert_sr_float(xls(sr_tab, row + 59, co("E")))
            s['yield_from_conv_practice'] = link_vma(sr_tab, row + 60, co("E"))

            assert xls(sr_tab, row + 72, co("B")) == 'Solution'
            assert xls(sr_tab, row + 73, co("D")) == 'First Cost:'
            s['pds_2014_cost'] = s['ref_2014_cost'] = link_vma(sr_tab, row + 73, co("E"))
            s['soln_first_cost_efficiency_rate'] = 0.0  # always 0 for LAND models
            s['soln_fixed_oper_cost_per_iunit'] = link_vma(sr_tab, row + 74, co("E"))
            s['soln_expected_lifetime'] = convert_sr_float(xls(sr_tab, row + 77, co("E")))
            s['yield_gain_from_conv_to_soln'] = link_vma(sr_tab, row + 78, co("E"))

            assert xls(sr_tab, row + 90, co("B")) == 'General'
            s['npv_discount_rate'] = convert_sr_float(xls(sr_tab, row + 91, co("E")))

            assert xls(sr_tab, row + 156, co("B")) == 'General Emissions Inputs'
            s['emissions_use_co2eq'] = convert_bool(xls(sr_tab, row + 157, co("E")))
            s['emissions_use_agg_co2eq'] = convert_bool(xls(sr_tab, row + 158, co("E")))
            s['emissions_grid_source'] = xls(sr_tab, row + 159, co("E"))
            s['emissions_grid_range'] = xls(sr_tab, row + 160, co("E"))

            assert xls(sr_tab, row + 144, co("B")) == 'Indirect Emissions'
            s['conv_indirect_co2_per_unit'] = convert_sr_float(sr_tab, row + 145, co("E"))
            s['soln_indirect_co2_per_iunit'] = convert_sr_float(sr_tab, row + 146, co("E"))

            assert xls(sr_tab, row + 132, co("B")) == 'Direct Emissions'
            s['tco2eq_reduced_per_land_unit'] = link_vma(sr_tab, row + 133, co("E"))
            s['tco2eq_rplu_rate'] = xls(sr_tab, row + 133, co("H"))
            s['tco2_reduced_per_land_unit'] = link_vma(sr_tab, row + 134, co("E"))
            s['tco2_rplu_rate'] = xls(sr_tab, row + 134, co("H"))
            s['tn2o_co2_reduced_per_land_unit'] = link_vma(sr_tab, row + 135, co("E"))
            s['tn2o_co2_rplu_rate'] = xls(sr_tab, row + 135, co("H"))
            s['tch4_co2_reduced_per_land_unit'] = link_vma(sr_tab, row + 136, co("E"))
            s['tch4_co2_rplu_rate'] = xls(sr_tab, row + 136, co("H"))
            s['land_annual_emissons_lifetime'] = convert_sr_float(sr_tab, row + 137, co("E"))

            assert xls(sr_tab, row + 109, co("B")) == 'Grid Emissions'
            s['conv_annual_energy_used'] = convert_sr_float(sr_tab, row + 110, co("E"))
            s['soln_annual_energy_used'] = convert_sr_float(sr_tab, row + 112, co("E"))

            assert xls(sr_tab, row + 168, co("B")) == 'Carbon Sequestration and Land Inputs'
            if xls(sr_tab, row + 169, co("E")) == "":
                # Excel checks whether this cell == "" to trigger different handling. The best equivalent
                # in Python is to set it to NaN. We can distinguish None (not set) from NaN, and if
                # the value is ever inadvertantly used it will result in NaN.
                s['seq_rate_global'] = np.nan

                if 'Variable Meta-analysis-DD' not in wb.sheetnames:
                    assert NotImplementedError(
                        'VMA Thermal-Moisture Regime sequestration not implemented')
                    # (4/2019) vma.py does have support for regimes in avg_high_low, it needs to be
                    # implemented in advanced_controls to pass a regime name through to vma.py

                tmr6 = tmr8 = False
                aez_tab = wb['AEZ Data']
                for x in range(aez_tab.min_row, aez_tab.max_row):
                    if aez_tab.cell(x, co("A")).value == 'TOTAL Boreal-Humid land':
                        tmr8 = True
                        break
                    if aez_tab.cell(x, co("A")).value == 'TOTAL Temperate/Boreal-Humid land':
                        tmr6 = True
                        break

                # For the public models using 'Variable Meta-analysis-DD', the DD tab does not contain
                # avg/high/low for the Thermal Moisture Regimes so we extract value from ScenarioRecord.
                if tmr6:
                    s['seq_rate_per_regime'] = {
                        'Tropical-Humid': convert_sr_float(sr_tab, row + 170, co("E")),
                        'Temperate/Boreal-Humid': convert_sr_float(sr_tab, row + 171, co("E")),
                        'Tropical-Semi-Arid': convert_sr_float(sr_tab, row + 172, co("E")),
                        'Temperate/Boreal-Semi-Arid': convert_sr_float(sr_tab, row + 173, co("E")),
                        'Global Arid': convert_sr_float(sr_tab, row + 174, 7),
                        'Global Arctic': 0.0}
                if tmr8:
                    s['seq_rate_per_regime'] = {
                        'Tropical-Humid': convert_sr_float(sr_tab, row + 170, co("E")),
                        'Temperate-Humid': convert_sr_float(sr_tab, row + 171, co("E")),
                        'Boreal-Humid': convert_sr_float(sr_tab, row + 171, co("E")),
                        'Tropical-Semi-Arid': convert_sr_float(sr_tab, row + 172, co("E")),
                        'Temperate-Semi-Arid': convert_sr_float(sr_tab, row + 173, co("E")),
                        'Boreal-Semi-Arid': convert_sr_float(sr_tab, row + 173, co("E")),
                        'Global Arid': convert_sr_float(sr_tab, row + 174, 7),
                        'Global Arctic': 0.0}
            else:
                s['seq_rate_global'] = link_vma(sr_tab, row + 169, co("E"))
            
            
            if xls(sr_tab, row + 175, co("D")) == 'Growth Rate of Land Degradation':
                s['global_multi_for_regrowth'] = convert_sr_float(sr_tab, row + 178, co("E"))
                s['degradation_rate'] = link_vma(sr_tab, row + 175, co("E"))
                s['tC_storage_in_protected_land_type'] = link_vma(sr_tab, row + 177, co("E"))
            elif xls(sr_tab, row + 175, co("D")) == 'Sequestered Carbon NOT Emitted after Cyclical Harvesting/Clearing':
                s['carbon_not_emitted_after_harvesting'] = link_vma(sr_tab, row + 175, co("E"))

            s['disturbance_rate'] = link_vma(sr_tab, row + 176, co("E"))

            assert xls(sr_tab, row + 188, co("B")) == 'General Land Inputs'
            if xls(sr_tab, row + 189, co("D")) == 'Delay Impact of Protection by 1 Year? (Leakage)':
                s['delay_protection_1yr'] = convert_bool(xls(sr_tab, row + 189, co("E")))
                s['delay_regrowth_1yr'] = convert_bool(xls(sr_tab, row + 190, co("E")))
                s['include_unprotected_land_in_regrowth_calcs'] = convert_bool(xls(sr_tab, row + 191, co("E")))
            elif xls(sr_tab, row + 189, co("D")) == 'New Growth is Harvested/Cleared Every':
                s['harvest_frequency'] = convert_sr_float(sr_tab, row + 189, co("E"))

            for addl in range(271, 285):
                if xls(sr_tab, row + addl, co("D")) == 'Avoided Deforested Area With Increase in Agricultural Intensification':
                    s['avoided_deforest_with_intensification'] = convert_sr_float(sr_tab.cell(row + addl, co("E")).value)
            scenarios[scenario_name] = s
    return scenarios





def normalize_source_name(sourcename):
    """Return a common name for widely used studies.
       Correct mis-spelings and inconsistencies in the column names.
       A few solution files, notably OnshoreWind, are inconsistent about the ordering
       of columns between different regions, and additionally use inconsistent names
       and some spelling errors. We need the column names to be consistent to properly
       combine them.

       World
       +-------------------+-------------------+-----------------------------+---------+
       |   Baseline Cases  |  Ambitious Cases  |     Conservative Cases      |100% REN |
       +-------------------+-------------------+-----------------------------+---------+
       | source1 | source2 | source3 | source4 | source5 | source6 | source7 | source8 |

       OECD90
       +-------------------+-------------------+-----------------------------+---------+
       |   Baseline Cases  |  Ambitious Cases  |     Conservative Cases      |100% REN |
       +-------------------+-------------------+-----------------------------+---------+
       | source2 | source1 | source3 | Source 4| surce5  | source6 | source7 | source8 |
    """
    special_cases = {
        'Based on: Greenpeace (2015) Reference': 'Based on: Greenpeace 2015 Reference',
        'Greenpeace 2015 Reference Scenario': 'Based on: Greenpeace 2015 Reference',
        'Based on Greenpeace 2015 Reference Scenario': 'Based on: Greenpeace 2015 Reference',
        'Based on: Greenpeace 2015 Reference Scenario': 'Based on: Greenpeace 2015 Reference',
        'Based on Greenpeace 2015 Reference': 'Based on: Greenpeace 2015 Reference',
        'Based on: Greenpeace Reference Scenario': 'Based on: Greenpeace 2015 Reference',
        'Conservative: Based on- Greenpeace 2015 Reference': 'Based on: Greenpeace 2015 Reference',
        '100% REN: Based on- Greenpeace Advanced [R]evolution': 'Based on: Greenpeace 2015 Advanced Revolution',
        'Drawdown TAM: Baseline Cases': 'Baseline Cases',
        'Drawdown TAM: Conservative Cases': 'Conservative Cases',
        'Drawdown TAM: Ambitious Cases': 'Ambitious Cases',
        'Drawdown TAM: Maximum Cases': 'Maximum Cases',
        'Drawdown Projections based on adjusted IEA data (ETP 2012) on projected growth in each year, and recent sales Data (IEA - ETP 2016)': 'Drawdown Projections based on adjusted IEA data (ETP 2012) on projected growth in each year, and recent sales Data (IEA - ETP 2016)',
        'ITDP/UC Davis (2014)  A Global High Shift Scenario Updated Report Data - Baseline Scenario': 'ITDP/UC Davis 2014 Global High Shift Baseline',
        'ITDP/UC Davis (2014)  A Global High Shift Scenario Updated Report Data - HighShift Scenario': 'ITDP/UC Davis 2014 Global High Shift HighShift',
        'What a Waste: A Global Review of Solid Waste Management (Hoornweg, 2012) - Static % of Organic Waste': 'What a Waste Solid Waste Management Static',
        'What a Waste: A Global Review of Solid Waste Management (Hoornweg, 2012) - Dynamic % of Organic Waste': 'What a Waste Solid Waste Management Dynamic',
        'What a Waste: A Global Review of Solid Waste Management (Hoornweg, 2012) - Dynamic Organic Fraction by Un Mediam Variant': 'What a Waste Solid Waste Management Dynamic Organic Fraction',
        'IPCC, 2006 - Calculated': 'IPCC, 2006 Calculated',
        "Combined from IEA (2016) ETP 2016, ICAO (2014) Annual Report 2014, Appendix 1, Boeing (2013) World Air cargo Forecast 2014-2015, Airbus (2014) Global market Forecast: Flying by the Numbers 2015-2034 - Highest Ranges": 'Combined from IEA ETP 2016, ICAO 2014, Boeing 2013, Airbus 2014, Highest Ranges',
        "Combined from IEA (2016) ETP 2016, ICAO (2014) Annual Report 2014, Appendix 1, Boeing (2013) World Air cargo Forecast 2014-2015, Airbus (2014) Global market Forecast: Flying by the Numbers 2015-2034 - Middle Ranges": 'Combined from IEA ETP 2016, ICAO 2014, Boeing 2013, Airbus 2014, Middle Ranges',
        "Combined from IEA (2016) ETP 2016, ICAO (2014) Annual Report 2014, Appendix 1, Boeing (2013) World Air cargo Forecast 2014-2015, Airbus (2014) Global market Forecast: Flying by the Numbers 2015-2034 - Lowest Ranges": 'Combined from IEA ETP 2016, ICAO 2014, Boeing 2013, Airbus 2014, Lowest Ranges',
        'Based on average of: LUT/EWG (2019) -100% RES; Ecofys (2018) - 1.5ºC and Greenpeace (2015) Advanced [R]evolution': 'Based on average of: LUT/EWG 2019 100% RES, Ecofys 2018 1.5C and Greenpeace 2015 Advanced Revolution',
        'FAO 2015 (Sum of all regions)': 'FAO 2015',  # Afforestation Drawdown 2020
        'FAO 2010 (Sum of all regions)': 'FAO 2010',  # Bamboo Drawdown 2020
    }
    if not sourcename:  # don't do anything with empty data
        return sourcename
    normalized = sourcename.replace("'", "").replace('\n', ' ').strip()
    if normalized in special_cases:
        return special_cases[normalized]
    if re.search(r'\[Source \d+', sourcename):
        return None

    # handle duplicate column names where xlrd appends an integer.
    suffix = ''
    r = re.search(r'(\\.\d)$', sourcename)
    if r is not None:
        suffix = r.group()

    name = re.sub(r"[\[\]]", "", sourcename.upper())  # [R]evolution to REVOLUTION
    if 'UN CES' in name and 'ITU' in name and 'AMPERE' in name:
        if 'BASELINE' in name: return 'Based on: CES ITU AMPERE Baseline' + suffix
        if '550' in name: return 'Based on: CES ITU AMPERE 550' + suffix
        if '450' in name: return 'Based on: CES ITU AMPERE 450' + suffix
        raise ValueError('Unknown UN CES ITU AMPERE source: ' + sourcename)
    if 'IEA' in name and 'ETP' in name:
        if '2014' in name and '2DS' in name: return 'Based on: IEA ETP 2014 2DS' + suffix
        if '2014' in name and '4DS' in name: return 'Based on: IEA ETP 2014 4DS' + suffix
        if '2014' in name and '6DS' in name: return 'Based on: IEA ETP 2014 6DS' + suffix
        if '2016' in name and '6DS' in name: return 'Based on: IEA ETP 2016 6DS' + suffix
        if 'ETP16' in name and '6DS' in name: return 'Based on: IEA ETP 2016 6DS' + suffix
        if '2016' in name and '4DS' in name: return 'Based on: IEA ETP 2016 4DS' + suffix
        if 'ETP16' in name and '4DS' in name: return 'Based on: IEA ETP 2016 4DS' + suffix
        if '2016' in name and '2DS' in name and 'OPT2-PERENNIALS' in name:
            return 'Based on: IEA ETP 2016 2DS with OPT2 perennials' + suffix
        if '2016' in name and '2DS' in name: return 'Based on: IEA ETP 2016 2DS' + suffix
        if 'ETP16' in name and '2DS' in name: return 'Based on: IEA ETP 2016 2DS' + suffix
        if '2016' in name and 'ANNEX' in name: return 'Based on: IEA ETP 2016 Annex' + suffix
        if '2017' in name and 'REF' in name: return 'Based on: IEA ETP 2017 Ref Tech' + suffix
        if '2017' in name and 'B2DS' in name: return 'Based on: IEA ETP 2017 B2DS' + suffix
        if '2017' in name and 'BEYOND 2DS' in name: return 'Based on: IEA ETP 2017 B2DS' + suffix
        if '2017' in name and '2DS' in name: return 'Based on: IEA ETP 2017 2DS' + suffix
        if '2017' in name and '4DS' in name: return 'Based on: IEA ETP 2017 4DS' + suffix
        if '2017' in name and '6DS' in name: return 'Based on: IEA ETP 2017 6DS' + suffix
        raise ValueError('Unknown IEA ETP source: ' + sourcename)
    if 'AMPERE' in name and 'MESSAGE' in name:
        if '450' in name: return 'Based on: AMPERE 2014 MESSAGE MACRO 450' + suffix
        if '550' in name: return 'Based on: AMPERE 2014 MESSAGE MACRO 550' + suffix
        if 'REF' in name: return 'Based on: AMPERE 2014 MESSAGE MACRO Reference' + suffix
        raise ValueError('Unknown AMPERE MESSAGE-MACRO source: ' + sourcename)
    if 'AMPERE' in name and 'IMAGE' in name:
        if '450' in name: return 'Based on: AMPERE 2014 IMAGE TIMER 450' + suffix
        if '550' in name: return 'Based on: AMPERE 2014 IMAGE TIMER 550' + suffix
        if 'REF' in name: return 'Based on: AMPERE 2014 IMAGE TIMER Reference' + suffix
        raise ValueError('Unknown AMPERE IMAGE-TIMER source: ' + sourcename)
    if 'AMPERE' in name and 'GEM' in name and 'E3' in name:
        if '450' in name: return 'Based on: AMPERE 2014 GEM E3 450' + suffix
        if '550' in name: return 'Based on: AMPERE 2014 GEM E3 550' + suffix
        if 'REF' in name: return 'Based on: AMPERE 2014 GEM E3 Reference' + suffix
        raise ValueError('Unknown AMPERE GEM E3 source: ' + sourcename)
    if 'GREENPEACE' in name and 'ENERGY' in name:
        if 'ADVANCED' in name and 'DRAWDOWN-PERENNIALS' in name:
            return 'Based on: Greenpeace 2015 Advanced Revolution with Drawdown perennials' + suffix
        if 'ADVANCED' in name: return 'Based on: Greenpeace 2015 Advanced Revolution' + suffix
        if 'REVOLUTION' in name and 'DRAWDOWN-PERENNIALS' in name:
            return 'Based on: Greenpeace 2015 Energy Revolution with Drawdown perennials' + suffix
        if 'REVOLUTION' in name: return 'Based on: Greenpeace 2015 Energy Revolution' + suffix
        if 'REFERENCE' in name: return 'Based on: Greenpeace 2015 Reference' + suffix
        raise ValueError('Unknown Greenpeace Energy source: ' + sourcename)
    if 'GREENPEACE' in name and 'THERMAL' in name:
        if 'MODERATE' in name: return 'Based on: Greenpeace 2016 Solar Thermal Moderate' + suffix
        if 'ADVANCED' in name: return 'Based on: Greenpeace 2016 Solar Thermal Advanced' + suffix
        raise ValueError('Unknown Greenpeace Solar Thermal source: ' + sourcename)
    return unicodedata.normalize('NFC', normalized)

def normalize_case_name(name):
    rewrites = {
        'Drawdown TAM: Baseline Cases': 'Baseline Cases',
        'Drawdown TAM: Conservative Cases': 'Conservative Cases',
        'Drawdown TAM: Ambitious Cases': 'Ambitious Cases',
        'Drawdown TAM: Maximum Cases': 'Maximum Cases',
        '100% Case': '100% RES2050 Case',
    }
    return rewrites.get(name, name)

def normalize_unit(tab, row, col):
    unit_mapping = {
        'Million hectare': u'Mha',
        'MMt FlyAsh Cement (Sol) or MMt OPC (Conv) (Transient)': u'MMt',
        'Billion USD': u'US$B',
        'million m2 commercial floor space': u'Mm\u00B2',
        'Million Households': u'MHholds',
        'Million m2 of Comm.+Resid. Floor Area Equiv. for Cold Climates': u'Mm\u00B2',
        'Giga-Liter Water': u'GL H\u2082O',
        'Million Metric tonnes per year': 'MMt',
        'million tonne-km': 'Mt-km',
        'million tonne-kms': 'Mt-km',
        'Residential and Commercial roof area, m2': u'm\u00B2',
        'Residential and Commercial roof area,  m2': u'm\u00B2',
    }
    name = xls(tab, row, col)
    return unit_mapping.get(name, name)




def get_filename_for_source(sourcename, prefix=''):
    """Return string to use for the filename for known sources."""
    if re.search(r'\[Source \d+', sourcename):
        return None
    if re.search(r'Drawdown TAM: \[Source \d+', sourcename):
        return None

    filename = sourcename.strip()
    filename = re.sub(r"[^\w\s\.]", '', filename)
    filename = re.sub(r"\s+", '_', filename)
    filename = re.sub(r"\.+", '_', filename)
    filename = filename.replace('Based_on_', 'based_on_')
    if len(filename) > 96:
        h = hashlib.sha256(filename.encode('utf-8')).hexdigest()[-8:]
        filename = filename[:88] + '_' + h
    return prefix + filename + '.csv'


def write_tam(f, wb, outputdir, is_elecgen=False):
    """Generate the TAM section of a solution.
       Arguments:
         f - file-like object for output
         wb - an Excel workbook as returned by openpyxl
         outputdir: name of directory to write CSV files to.
         is_elecgen: True if this is an electricity generation model
    """

    tm_tab = wb['TAM Data']
    lk = lambda x : xls(tm_tab, x)
 
    f.write( "\n")
    f.write( "        # Instructions: Set TAM override parameters appropriately if any of these vary from the standard (then delete these comments):\n")
    f.write(f"        # trend (3rd Poly): {lk('B19')} {lk('L17')} {lk('L20')} {lk('L23')} {lk('L26')} {lk('L29')} {lk('L32')} {lk('L35')} {lk('L38')} {lk('L41')}\n")
    f.write(f"        # growth (medium): {lk('C19')} {lk('M17')} {lk('M20')} {lk('M23')} {lk('M26')} {lk('M29')} {lk('M32')} {lk('M35')} {lk('M38')} {lk('M41')}\n")
    f.write(f"        # low_sd_mult (1.0): {lk('B25')} {lk('Q17')} {lk('Q20')} {lk('Q23')} {lk('Q26')} {lk('Q29')} {lk('Q32')} {lk('Q35')} {lk('Q38')} {lk('Q41')}\n")
    f.write(f"        # high_sd_mult (1.0): {lk('B24')} {lk('Q16')} {lk('Q19')} {lk('Q22')} {lk('Q25')} {lk('Q28')} {lk('Q31')} {lk('Q34')} {lk('Q37')} {lk('Q40')}\n") 
    f.write( "\n")

    if is_elecgen:  # Special case energy solutions, because they use shared data sources.
        f.write("        self._ref_tam_sources = scenario.load_sources(rrs.energy_ref_tam(),'*')\n")
        f.write("        self._pds_tam_sources = scenario.load_sources(rrs.energy_pds_tam(),'*')\n")
    else:
        # Is it really stupid to write comments in the code about bugs, rather than fixing the bugs?
        # Yes, yes, it is.
        f.write( "        # Also, there is currently a bug in the code that generates the TAM json file, which\n")
        f.write( "        # causes some sources to be assigned to the empty case.  Please check if this is\n")
        f.write( "        # occurring, and if so, put the sources in the case where they belong.\n")
        f.write( "\n")

        # Extract the data sources
        # See extract_source_data for the definition of these line numbers
        tam_regions = {'World': 44, 'OECD90': 162, 'Eastern Europe': 226,
                    'Asia (Sans Japan)': 289, 'Middle East and Africa': 352, 'Latin America': 415,
                    'China': 478, 'India': 542, 'EU': 606, 'USA': 671}
        tamoutputdir = os.path.join(outputdir, 'tam')
        os.makedirs(tamoutputdir, exist_ok=True)
        ref_sources = extract_source_data(wb=wb, sheet_name='TAM Data', regions=tam_regions,
                                        outputdir=tamoutputdir, prefix='tam_')

        write_json(filename=Path(tamoutputdir)/'tam_ref_sources.json', d=ref_sources)
        f.write("        self._ref_tam_sources = scenario.load_sources(THISDIR/'tam/tam_ref_sources.json','*')\n")

        tam_regions = {'World': 102}
        pds_sources = extract_source_data(wb=wb, sheet_name='TAM Data', regions=tam_regions,
                                        outputdir=tamoutputdir, prefix='tam_pds_')
        if pds_sources:
            write_json(filename=Path(tamoutputdir)/'tam_pds_sources.json', d=ref_sources)
            f.write("        self._pds_tam_sources = scenario.load_sources(THISDIR/'tam/tam_pds_sources.json','*')\n")
        else:
            f.write("        self._pds_tam_sources = self._ref_tam_sources\n")

    regional = convert_bool(xls(tm_tab, 'B29')) and convert_bool(xls(tm_tab, 'B30'))
    f.write(f"        self.set_tam({'main_includes_regional=True' if regional else ''})\n")
    f.write( "        ref_tam_per_region=self.tm.ref_tam_per_region()\n")
    f.write( "        pds_tam_per_region=self.tm.pds_tam_per_region()\n")
    f.write("\n")


def write_aez(f, wb, use_custom_tla):
    a = wb['Land Allocation - Max TLA']
    first_solution = xls(a, 'B18')
    if first_solution == 'Peatland Protection':
        f.write("        self.ae = aez.AEZ(solution_name=self.name, cohort=2020,\n")
        f.write("                regimes=dd.THERMAL_MOISTURE_REGIMES8)\n")
    elif first_solution == 'Forest Protection':
        f.write("        self.ae = aez.AEZ(solution_name=self.name, cohort=2018)\n")
    else:
        raise ValueError('cannot determine AEZ Land Allocation to use')

    if use_custom_tla:
        f.write("        if self.ac.use_custom_tla and self.ac.custom_tla_fixed_value is not None:\n")
        f.write("            self.c_tla = tla.CustomTLA(fixed_value=self.ac.custom_tla_fixed_value)\n")
        f.write("            custom_world_vals = self.c_tla.get_world_values()\n")
        f.write("        elif self.ac.use_custom_tla:\n")
        f.write("            self.c_tla = tla.CustomTLA(filename=THISDIR/'custom_tla_data.csv')\n")
        f.write("            custom_world_vals = self.c_tla.get_world_values()\n")
        f.write("        else:\n")
        f.write("            custom_world_vals = None\n")
        f.write("        self.tla_per_region = tla.tla_per_region(self.ae.get_land_distribution(),\n")
        f.write("            custom_world_values=custom_world_vals)\n\n")
    else:
        f.write("        self.tla_per_region = tla.tla_per_region(self.ae.get_land_distribution())\n\n")


def write_ad(f, wb, outputdir):
    """Generate the Adoption Data section of a solution.
       Arguments:
         f - file-like object for output
         wb - an Excel workbook as returned by xlrd
         outputdir: name of directory to write CSV files to.
    """
    
    ad_regions = find_ad_regions(wb=wb)
    ad_outputdir = os.path.join(outputdir, 'ad')
    sources = extract_source_data(wb=wb, sheet_name='Adoption Data', regions=ad_regions,
            outputdir=ad_outputdir, prefix='ad_')

    if sources:
        a = wb['Adoption Data']
        lk = lambda x : xls(a, x)
        # Extract the fit parameters and output them for users to double check that they don't need
        # special handling
        f.write( "        # Instructions: Set AD override parameters appropriately if any of these regional values vary from the standard\n")
        f.write( "        # (then delete these comments):\n")
        f.write(f"        # trend (3rd Poly): {lk('L17')} {lk('L20')} {lk('L26')} {lk('L29')} {lk('L32')} {lk('L35')} {lk('L38')} {lk('L41')}\n")
        f.write(f"        # growth (medium): {lk('M17')} {lk('M20')} {lk('M23')} {lk('M26')} {lk('M29')} {lk('M32')} {lk('M35')} {lk('M38')} {lk('M41')}\n")
        if xls(a, 'R17') == 'S.D.':
            f.write(f"        # low_sd_mult (1.0): {lk('Q17')} {lk('Q20')} {lk('Q23')} {lk('Q26')} {lk('Q29')} {lk('Q32')} {lk('Q35')} {lk('Q38')} {lk('Q41')}\n")
            f.write(f"        # high_sd_mult (1.0): {lk('Q16')} {lk('Q19')} {lk('Q22')} {lk('Q25')} {lk('Q28')} {lk('Q31')} {lk('Q34')} {lk('Q37')} {lk('Q40')}\n") 

        regional = convert_bool(xls(a, 'B30')) and convert_bool(xls(a, 'B31'))
        if regional:
            f.write("        self._pds_ad_settings['main_includes_regional'] = True\n") 
        if is_elecgen:
            f.write("        # Quirks parameter should apply to energy solutions only (remove once ")
            f.write("        self._pds_ad_settings['groups_include_hundred_percent'] = False\n")
        f.write("        self._pds_ad_sources = scenario.load_sources(THISDIR/'ad/ad_sources.json', '*')\n")


def write_ca(case, f, wb, outputdir):
    """Generate the Custom Adoption Data section of a solution.
       Arguments:
         case: 'PDS' or 'REF'
         f: file-like object for output
         wb: an Excel workbook as returned by openpyxl
         outputdir: name of the solution directory
    """
    lcase = case.lower()
    prefix = f"ca_{lcase}"
    datadir = Path(outputdir)/f'{prefix}_data'
    scenarios, multipliers = extract_custom_adoption(wb=wb, outputdir=datadir,
                                                     sheet_name=f'Custom {case} Adoption',
                                                     prefix=prefix)
    if scenarios:
        if is_land:
            f.write( "       # CAUTION: Many Land solutions have highly customized custom adoption setups.\n")
            f.write( "       # Check older versions of this file, or similar solution types, to determine if\n")
            f.write( "       # this code must be replaced with completely custom code.\n") 
        if multipliers['high'] != 1.0 or multipliers['low'] != 1.0:
            f.write(f"        self._{lcase}_ca_settings = \{ 'hi_sd_mult': {multipliers['high']}, 'low_sd_mult': {multipliers['low']} \}\n")  
        f.write(f"        self._{lcase}_ca_sources = scenario.load_sources(THISDIR/'{prefix}_data/{prefix}_sources.json', 'filename')\n")    


def write_ht(f, wb):
    """Generate the Helper Tables section of a solution.
       Arguments:
         f: file-like object for output
         wb: an Excel workbook as returned by openpyxl
         has_single_source: whether to emit a pds_adoption_is_single_source arg
    """
    h = wb['Helper Tables']
    ref_initial_year = h['B21'].value
    pds_initial_year = h['B85'].value

    tam_or_tla = 'ref_tam_per_region' if not is_land else 'self.tla_per_region'
    f.write( "        final_year=2050  # Currently fixed for all models; may be variable in the future.\n")
    f.write( "        ht_ref_adoption_initial = pd.Series(self.ac.ref_base_adoption)\n")
    f.write(f"        ht_ref_adoption_final = ({tam_or_tla}.loc[final_year] * \n")
    f.write(f"            (ht_ref_adoption_initial / {tam_or_tla}.loc[self.base_year]))\n")
    f.write( "        ht_ref_datapoints = pd.DataFrame(columns=dd.REGIONS)\n")
    f.write( "        ht_ref_datapoints.loc[self.base_year] = ht_ref_adoption_initial\n")
    f.write(f"        ht_ref_datapoints.loc[final_year] = ht_ref_adoption_final\n")


    tam_or_tla = 'pds_tam_per_region' if not is_land else 'self.tla_per_region'
    f.write(f"        pds_initial_year = {pds_initial_year}  # sometimes, but rarely, different than self.base_year\n")
    f.write( "                                # Excel 'Helper Tables'!B85\n")
    f.write( "        ht_pds_adoption_initial = ht_ref_adoption_initial\n")
    f.write( "        ht_pds_adoption_final_percentage = pd.Series(self.ac.pds_adoption_final_percentage)\n")
    f.write(f"        ht_pds_adoption_final = ht_pds_adoption_final_percentage * {tam_or_tla}.loc[final_year]\n")
    f.write( "        ht_pds_datapoints = pd.DataFrame(columns=dd.REGIONS)\n")
    f.write( "        ht_pds_datapoints.loc[pds_initial_year] = ht_pds_adoption_initial\n")
    f.write( "        ht_pds_datapoints.loc[final_year] = ht_pds_adoption_final\n")

    f.write( "        self.ht = helpertables.HelperTables(ac=self.ac,\n")
    f.write( "            ref_datapoints=ht_ref_datapoints,\n")
    f.write( "            pds_datapoints=ht_pds_datapoints,\n")
    f.write( "            ref_adoption_data_per_region=ref_adoption_data_per_region,\n")
    f.write( "            pds_adoption_data_per_region=pds_adoption_data_per_region,\n")
    if is_land:
        f.write( "            ref_adoption_limits=self.tla_per_region,\n")
        f.write( "            pds_adoption_limits=self.tla_per_region,\n")
    else:
        f.write( "            ref_adoption_limits=ref_tam_per_region,\n")
        f.write( "            pds_adoption_limits=pds_tam_per_region,\n")
    f.write( "            pds_adoption_trend_per_region=pds_adoption_trend_per_region,\n")

    # Assess the Quirks Parameters
    # Note that we could do a more reliable job here by utilizing the fact that openpyxl can
    # read the formulas of the cells, not just their contents.  Just don't have time to do that now.

    v = lambda r, c: h.cell(r,c).value  # shortcut, and skip the data cleaning in this case

    # check if pds is being copied to ref
    # First check if there are any Y-PDS type values --- that would give us a false positive, and a circular
    # calculation to boot (discovered the hard way)
    copy_pds_to_ref = False
    ac = wb['Advanced Controls']
    start_row = find_in_column(ac, co("F"), "Adjustment?", 250)
    assert start_row, "Couldn't find Adjustment? column on AC sheet"
    y_pds = False
    for row in range(start_row+1,start_row+5):
        if xls(ac,row,co("F")) != "N":
            y_pds = True
            break

    # If there were no Y-PDS in the way, we detect copy_pds_to_ref by looking for shared values
    # in the two tables.
    if not y_pds:  
        offset = 0
        # we actually start at row 2015, because 2014 treatment sometimes has other issues
        while v(28+offset, co("C")) == v(92+offset, co("C")):
            offset += 1
        copy_pds_to_ref = (offset > 0)
        copy_through_year = 2014+offset

    # Now, lets check the behavior of the first rows of both tables
    # the default _is_ to copy, so we are looking for evidence that we should not
    copy_ref = True
    copy_ref_world_too = False
    base_row = 27 + (ref_initial_year - 2014)  # where would we be copying to?
    for col in range(co("D"),co("L")+1):
        if v(19,col) != v(base_row,col):
            copy_ref = False
            break
    if copy_ref and v(19,co("C")) == v(base_row,co("C")):
        copy_ref_world_too = True
    
    copy_pds = True
    copy_pds_world_too = False
    base_row = 91 + (pds_initial_year - 2014)
    for col in range(co("D"),co("L")+1):
        if v(85,col) != v(base_row,col):
            copy_pds = False
            break
    if copy_pds and v(85,co("C")) == v(base_row,co("C")):
        copy_pds_world_too = True

    f.write( "            # Quirks Parameters.  The generator tries to guess these correctly, but can get\n")
    f.write( "            # it wrong.  See the documentation for HelperTables.__init__() to understand\n")
    f.write( "            # exactly what the paramaters do, and how to set them.\n")
    
    f.write(f"            copy_pds_to_ref={copy_pds_to_ref},\n")
    if copy_pds_to_ref:
        f.write(f"            copy_through_year={copy_through_year},\n")
    f.write(f"            copy_ref_datapoint={copy_ref},\n")
    if copy_ref:
        f.write(f"            copy_ref_world_too={copy_ref_world_too},\n")
    f.write(f"            copy_pds_datapoint={copy_pds},\n")
    if copy_pds:
        f.write(f"            copy_pds_world_too={copy_pds_world_too},\n")

    f.write( "            pds_adoption_is_single_source=pds_adoption_is_single_source)\n")
    f.write( "\n")


def write_ef(f, wb):
    """Write out the Emissions Factors module for this solution class."""
    f.write("        # Emissions: if this is an older model, you may need to set a data version to make tests pass.\n")
    f.write("        self.ef = emissionsfactors.ElectricityGenOnGrid(ac=self.ac)\n")
    f.write("\n")


def write_ua(f, wb, is_rrs=True):
    """Write out the Unit Adoption module for this solution class."""
    ua_tab = wb['Unit Adoption Calculations']
    ac_tab = wb['Advanced Controls']
    f.write("        self.ua = unitadoption.UnitAdoption(ac=self.ac,\n")
    if is_rrs:
        f.write("            ref_total_adoption_units=ref_tam_per_region,\n")
        f.write("            pds_total_adoption_units=pds_tam_per_region,\n")
    else:
        f.write("            ref_total_adoption_units=self.tla_per_region,\n")
        f.write("            pds_total_adoption_units=self.tla_per_region,\n")
        f.write("            electricity_unit_factor=1000000.0,\n")
    f.write("            soln_ref_funits_adopted=self.ht.soln_ref_funits_adopted(),\n")
    f.write("            soln_pds_funits_adopted=self.ht.soln_pds_funits_adopted(),\n")
    if 'Repeated First Cost to Maintaining Implementation Units' in xls(ac_tab, 'A43'):
        repeated_cost_for_iunits = convert_bool(xls(ac_tab, 'C43'))
        f.write("            repeated_cost_for_iunits=" + str(repeated_cost_for_iunits) + ",\n")
    # If S135 == D135 (for all regions), then it must not be adding in 'Advanced Controls'!C62
    bug_cfunits_double_count = False
    for i in range(0, 9):
        if ua_tab.cell(135, 19 + i).value != ua_tab.cell(135, 4 + i).value:
            bug_cfunits_double_count = True
    f.write("            bug_cfunits_double_count=" + str(bug_cfunits_double_count) + ")\n")
    f.write("        soln_pds_tot_iunits_reqd = self.ua.soln_pds_tot_iunits_reqd()\n")
    f.write("        soln_ref_tot_iunits_reqd = self.ua.soln_ref_tot_iunits_reqd()\n")
    f.write("        conv_ref_tot_iunits = self.ua.conv_ref_tot_iunits()\n")
    f.write("        soln_net_annual_funits_adopted=self.ua.soln_net_annual_funits_adopted()\n")
    f.write("\n")


def write_fc(f, wb):
    """Code generate the First Code module for this solution class."""
    fc_tab = wb['First Cost']
    f.write("        self.fc = firstcost.FirstCost(ac=self.ac, pds_learning_increase_mult=" + xls(fc_tab, 'C25') + ",\n")
    f.write("            ref_learning_increase_mult=" + xls(fc_tab, 'D25') +
            ", conv_learning_increase_mult=" + xls(fc_tab, 'E25') + ",\n")
    f.write("            soln_pds_tot_iunits_reqd=soln_pds_tot_iunits_reqd,\n")
    f.write("            soln_ref_tot_iunits_reqd=soln_ref_tot_iunits_reqd,\n")
    f.write("            conv_ref_tot_iunits=conv_ref_tot_iunits,\n")
    f.write("            soln_pds_new_iunits_reqd=self.ua.soln_pds_new_iunits_reqd(),\n")
    f.write("            soln_ref_new_iunits_reqd=self.ua.soln_ref_new_iunits_reqd(),\n")
    f.write("            conv_ref_new_iunits=self.ua.conv_ref_new_iunits(),\n")
    if xls(fc_tab, 'P36') == 'Implementation Units Installed Each Yr (CONVENTIONAL-REF)':
        f.write("            conv_ref_first_cost_uses_tot_units=True,\n")
    if xli(fc_tab, 'F15') == 1000000000 and xls(fc_tab, 'G15') == '$/kW TO $/TW':
        f.write("            fc_convert_iunit_factor=conversions.terawatt_to_kilowatt())\n")
    elif xli(fc_tab, 'F16') == 1000000 and xls(fc_tab, 'F18') == 'million hectare':
        f.write("            fc_convert_iunit_factor=conversions.mha_to_ha)\n")
    else:
        f.write("            fc_convert_iunit_factor=" + xls(fc_tab, 'F15') + ")\n")
    f.write('\n')


def write_oc(f, wb):
    """Code generate the Operating Code module for this solution class."""
    oc_tab = wb['Operating Cost']
    f.write("        self.oc = operatingcost.OperatingCost(ac=self.ac,\n")
    f.write("            soln_net_annual_funits_adopted=soln_net_annual_funits_adopted,\n")
    f.write("            soln_pds_tot_iunits_reqd=soln_pds_tot_iunits_reqd,\n")
    f.write("            soln_ref_tot_iunits_reqd=soln_ref_tot_iunits_reqd,\n")
    f.write("            conv_ref_annual_tot_iunits=self.ua.conv_ref_annual_tot_iunits(),\n")
    f.write("            soln_pds_annual_world_first_cost=self.fc.soln_pds_annual_world_first_cost(),\n")
    f.write("            soln_ref_annual_world_first_cost=self.fc.soln_ref_annual_world_first_cost(),\n")
    f.write("            conv_ref_annual_world_first_cost=self.fc.conv_ref_annual_world_first_cost(),\n")
    f.write("            single_iunit_purchase_year=" + xls(oc_tab, 'I121') + ",\n")
    f.write("            soln_pds_install_cost_per_iunit=self.fc.soln_pds_install_cost_per_iunit(),\n")
    f.write("            conv_ref_install_cost_per_iunit=self.fc.conv_ref_install_cost_per_iunit(),\n")

    units = xls(oc_tab, 'F13')
    is_energy_units = (units == '$/kW TO $/TW' or units == 'From US$2014 per kW to US$2014 per TW')
    conversion_factor_fom = xls(oc_tab,'E13')
    conversion_factor_vom = xls(oc_tab,'E14')

    if conversion_factor_fom == '1000000000' and is_energy_units:
        conversion_factor_fom = 'conversions.terawatt_to_kilowatt()'
    if conversion_factor_vom == '1000000000' and is_energy_units:
        conversion_factor_vom = 'conversions.terawatt_to_kilowatt()'
    if is_land:
        conversion_factor_fom = conversion_factor_vom = 'conversions.mha_to_ha()'

    # In almost all cases the two conversion factors are equal. We only know of one solution where
    # they differ (Heatpumps). operatingcost.py accomodates this, if passed a single number it will
    # use it for both factors.
    if conversion_factor_fom == conversion_factor_vom:
        f.write("            conversion_factor=" + conversion_factor_fom + ")\n")
    else:
        f.write("            conversion_factor=(" + conversion_factor_fom + ", " +
                conversion_factor_vom + "))\n")
    f.write('\n')


def write_c2_c4(f, is_protect=False, has_harvest=False):
    """Write out the CO2 Calcs and CH4 Calcs modules for this solution class."""
    f.write("        self.c4 = ch4calcs.CH4Calcs(ac=self.ac,\n")
    if not is_rrs:
        f.write("            soln_pds_direct_ch4_co2_emissions_saved=self.ua.direct_ch4_co2_emissions_saved_land(),\n")
    f.write("            soln_net_annual_funits_adopted=soln_net_annual_funits_adopted)\n\n")
    f.write("        self.c2 = co2calcs.CO2Calcs(ac=self.ac,\n")
    f.write("            ch4_ppb_calculator=self.c4.ch4_ppb_calculator(),\n")
    f.write("            soln_pds_net_grid_electricity_units_saved=self.ua.soln_pds_net_grid_electricity_units_saved(),\n")
    f.write("            soln_pds_net_grid_electricity_units_used=self.ua.soln_pds_net_grid_electricity_units_used(),\n")
    if is_rrs:
        f.write("            soln_pds_direct_co2_emissions_saved=self.ua.soln_pds_direct_co2_emissions_saved(),\n")
        f.write("            soln_pds_direct_ch4_co2_emissions_saved=self.ua.soln_pds_direct_ch4_co2_emissions_saved(),\n")
        f.write("            soln_pds_direct_n2o_co2_emissions_saved=self.ua.soln_pds_direct_n2o_co2_emissions_saved(),\n")
    else:
        f.write("            soln_pds_direct_co2eq_emissions_saved=self.ua.direct_co2eq_emissions_saved_land(),\n")
        f.write("            soln_pds_direct_co2_emissions_saved=self.ua.direct_co2_emissions_saved_land(),\n")
        f.write("            soln_pds_direct_n2o_co2_emissions_saved=self.ua.direct_n2o_co2_emissions_saved_land(),\n")
        f.write("            soln_pds_direct_ch4_co2_emissions_saved=self.ua.direct_ch4_co2_emissions_saved_land(),\n")
    f.write("            soln_pds_new_iunits_reqd=self.ua.soln_pds_new_iunits_reqd(),\n")
    f.write("            soln_ref_new_iunits_reqd=self.ua.soln_ref_new_iunits_reqd(),\n")
    f.write("            conv_ref_new_iunits=self.ua.conv_ref_new_iunits(),\n")
    f.write("            conv_ref_grid_CO2_per_KWh=self.ef.conv_ref_grid_CO2_per_KWh(),\n")
    f.write("            conv_ref_grid_CO2eq_per_KWh=self.ef.conv_ref_grid_CO2eq_per_KWh(),\n")
    f.write("            soln_net_annual_funits_adopted=soln_net_annual_funits_adopted,\n")
    if is_rrs:
        f.write("            fuel_in_liters=False)\n")
    else:
        if is_protect:
            f.write("            tot_red_in_deg_land=self.ua.cumulative_reduction_in_total_degraded_land(),\n")
            f.write("            pds_protected_deg_land=self.ua.pds_cumulative_degraded_land_protected(),\n")
            f.write("            ref_protected_deg_land=self.ua.ref_cumulative_degraded_land_protected(),\n")
        if has_harvest:
            f.write("            annual_land_area_harvested=self.ua.soln_pds_annual_land_area_harvested(),\n")
        f.write("            regime_distribution=self.ae.get_land_distribution(),\n")
        f.write("            regimes=dd.THERMAL_MOISTURE_REGIMES8)\n")
    f.write("\n")


def find_source_data_columns(wb, sheet_name, row):
    """Figure out which columns in Adoption Data (and similar tabs) should be extracted.
       Arguments:
         wb: Excel workbook
         sheet_name: name of the spreadsheet tab like "Adoption Data" or "Fresh Adoption Data"
         row: row number to check
       Returns:
         The string of Excel columns to use, like 'B:R'
    """
    tab = wb[sheet_name]
    for col in range(2, tab.max_column):
        # Look one row down for 'Functional Unit' as a stopping condition
        if tab.cell(row+1, col).value == 'Functional Unit':
            break
    return 'B:' + openpyxl.utils.cell.get_column_letter(col-1)


def data_sources_equivalent_for_region(region, world):
    for case, region_sources in region.items():
        world_sources = world.get(case, {})
        for source, region_filename in region_sources.items():
            world_filename = world_sources.get(source, None)
            if region_filename != world_filename:
                return False
    return True


def find_ad_regions(wb):
    # There are several different layouts for adoption data in different spreadsheets.  See if one
    # of the variants is detected.
    ad_default = {'World': 44, 'OECD90': 104, 'Eastern Europe': 168, 'Asia (Sans Japan)': 231,
                  'Middle East and Africa': 294, 'Latin America': 357, 'China': 420, 'India': 484,
                  'EU': 548, 'USA': 613}
    ad_microwind = {'World': 44, 'OECD90': 296, 'Eastern Europe': 170, 'Asia (Sans Japan)': 233,
                  'Middle East and Africa': 359, 'Latin America': 423, 'China': 487, 'India': 613,
                  'EU': 107, 'USA': 552}
    ad_afforestation = {'World': 44, 'OECD90': 102, 'Eastern Europe': 164, 'Asia (Sans Japan)': 225,
                  'Middle East and Africa': 286, 'Latin America': 347, 'China': 408, 'India': 470,
                  'EU': 530, 'USA': 591}
    tab = wb["Adoption Data"]
    for candidate in [ad_microwind, ad_afforestation]:
        for region, row in candidate.items():
            if region == 'World':
                continue
            found = False
            # Look for the name in the 'A' column anywhere slightly above or slightly below
            for r in range(row - 10, row+2):
                if region.lower() in xls(tab, r, co("A")).lower():
                    found = True
                    break
            if not found:
                break
        else:
            return candidate
    return ad_default


def extract_source_data(wb, sheet_name, regions, outputdir, prefix):
    """Pull the names of sources, by case, from the Excel file and write data to CSV.
       Arguments:
         wb: Excel workbook
         sheet_name: name of the Excel tab to go to, like 'Adoption Data' or 'TAM Data'
         regions: a dict of regions to extract and the line numbers they start at:
           { 'World': 44, 'OECD90': 104, 'Eastern Europe': 168 ...}
           The starting line is the 1-based index of the line with the Baseline/Ambitious etc. header
         outputdir: name of directory to write CSV files to.
         prefix: prefix for filenames like 'ad_' or 'tam_'

       Returns: a dict of category keys, containing lists of source names.

       Parsing the data sources can be messy: the number of sources within each category
       varies between solutions, and even the names of the cases can vary. Most solutions have
       Baseline/Ambitious/Conservative, but we're not sure that all do, and the final most
       aggressive case varies in names like "100% REN" (renewables) or "Maximum Cases".
       +-------------------+-------------------+-----------------------------+---------+
       |   Baseline Cases  |  Ambitious Cases  |     Conservative Cases      |100% REN |
       +-------------------+-------------------+-----------------------------+---------+
       | source1 | source2 | source3 | source4 | source5 | source6 | source7 | source8 | Functional Unit

       The category label like "Baseline Cases" is a merged cell one, two, three, or more cells
       across. In xlrd, the first cell contains the string and the subsequent cells are ctype
       XL_CELL_EMPTY. They have a border, but at the time of this writing xlrd does not extract
       styling information like borders from xlsx/xlsm files (only the classic Excel file format).
    """

    # Read the entire dataset for each region, across all sources, starting at the sourcename row
    region_data = {}
    for (region, baserow) in regions.items():
        usecols = find_source_data_columns(wb=wb, sheet_name=sheet_name, row=baserow)
        # Here we're reading the table starting at the sourcename header.  Since pd.read_excel
        # wants offsets in 0-base, it works out to the same row number as the 1-based case-name header
        df = pd.read_excel(wb, engine='openpyxl', sheet_name=sheet_name, header=baserow,
                           index_col=0, usecols=usecols, nrows=49)
        df.name = region
        df.rename(columns=normalize_source_name, inplace=True)
        region_data[region] = df

    # Collect all the source names that appear across all regions (allows for regions
    # having different sources than other regions)
    sources = {}
    for df in region_data.values():
        for source_name in df.columns:
            if source_name is not None:
                sources[source_name] = ''
    
    # No sources, nothing to do here.
    if not sources:
        return {}

    outputdir = Path(outputdir)
    outputdir.mkdir(exist_ok=True)

    # Pivot the data to source-centric: for each source, create a df with a column for
    # each region (even if it is empty)
    for source_name in list(sources.keys()):
        if not source_name:
            continue
        df = pd.DataFrame()
        for region in region_data.keys():
            if source_name in region_data[region].columns:
                df[region] = region_data[region].loc[:, source_name]
            else:
                df[region] = np.nan
        filename = get_filename_for_source(source_name, prefix=prefix)
        if df.empty or df.isna().all(axis=None, skipna=False) or not filename:
            del sources[source_name]
            continue
        df.index = df.index.astype(int)
        df.index.name = 'Year'

        zero_adoption_ok = outputdir.parent.name in zero_adoption_solutions
        if not zero_adoption_ok:
            # In the Excel implementation, adoption data of 0.0 is treated the same as N/A,
            # no data available. We don't want to implement adoptiondata.py the same way, we
            # want to be able to express the difference between a solution which did not
            # exist prior to year N, and therefore had 0.0 adoption, from a solution which
            # did exist but for which we have no data prior to year N.
            # We're handling this here: when extracting adoption data from
            # an Excel file, treat values of 0.0 as N/A and write out a CSV file with no
            # data at that location.
            df.replace(to_replace=0.0, value=np.nan, inplace=True)

        outputfile = outputdir/filename
        df.to_csv(outputfile, header=True, encoding='utf-8')
        sources[source_name] = filename

    tmp_cases = {}
    tab = wb[sheet_name]
    case_line = regions['World']
    for (region, line) in regions.items():
        case = ''
        for col in range(2, tab.max_column):
            if tab.cell(line+1, col).value == 'Functional Unit':
                break
            case = xls(tab, case_line, col)
            if case != '':
                case = normalize_case_name(case)
            # it is important to get the source name from the regional_data.columns here, not re-read
            # the source_name from Excel, because some solutions like Composting have duplicate
            # column names and pd.read_excel automatically appends ".1" and ".2" to make them unique.
            source_name = region_data[region].columns[col - 3]  
                 # col is the excel index, which is 1-based and has 2 columns to the left, so -3
                 # gives the equivalent df index.
            filename = sources.get(source_name, None)
            if source_name is not None and filename is not None:
                key = 'Region: ' + region
                region_cases = tmp_cases.get(key, dict())
                tmp_cases[key] = region_cases
                s = region_cases.get(case, dict())
                if source_name not in s:
                    s[source_name] = filename
                    region_cases[case] = s

    # suppress regions which are a subset of World.
    if 'Region: World' in tmp_cases:
        world = tmp_cases['Region: World']
        del tmp_cases['Region: World']
        cases = world.copy()
        for (region_name, sources) in tmp_cases.items():
            if not data_sources_equivalent_for_region(region=sources, world=world):
                cases[region_name] = sources
    else:
        cases = tmp_cases
    return cases


def extract_custom_adoption(wb, outputdir, sheet_name, prefix):
    """Extract custom adoption scenarios from an Excel file.
       Arguments:
         wb: Excel workbook as returned by openpyxl
         outputdir: directory where output files are written
         sheet_name: Excel sheet name to extract from
         prefix: string to prepend to filenames
    """
    custom_ad_tab = wb[sheet_name]

    assert xls(custom_ad_tab, 'AN25') == 'High'
    multipliers = {'high': xli(custom_ad_tab, 'AO25'),
                   'low': xli(custom_ad_tab, 'AO26')}
    scenarios = []
    # Look for the list of scenarios in the minitable in columns 'N:O'
    for srow in range(20, 36):
        if not re.search(r"Scenario \d+", xls(custom_ad_tab, srow, co("N"))):
            continue
        name = normalize_source_name(xls(custom_ad_tab, srow, co("O")))
        filename = get_filename_for_source(name, prefix=prefix+'_')
        if not filename:   # This skips all the "[Type scenario name here]" lines
            continue
        skip = True
        description = ''

        # Now find the main table for this scenario
        for row in range(custom_ad_tab.min_row, custom_ad_tab.max_row):
            if normalize_source_name(xls(custom_ad_tab, row, co("B"))) == name:
                # row points to 1-based top-level header, which is equiv to 0-based secondary header used by 
                # pd.read_excel used below.  i.e. this read starts at the 2nd header.
                df = pd.read_excel(wb, engine='openpyxl', sheet_name=sheet_name,
                                   header=row, index_col=0, usecols="A:K", nrows=49)
                df.rename(mapper={'Middle East & Africa': 'Middle East and Africa',
                          'Asia (sans Japan)': 'Asia (Sans Japan)'},
                          axis='columns', inplace=True)
                if not df.dropna(how='all', axis=1).dropna(how='all', axis=0).empty:
                    outputdir.mkdir(exist_ok=True)
                    df.to_csv(outputdir/filename, index=True, header=True, encoding='utf-8')
                    skip = False
                for offset in range(0, 3):
                    # TODO: deal with unicode on windows here.
                    # looking for the big yellow box.
                    description = xls(custom_ad_tab, row + offset, co("N"))
                    if description:
                        break
                break
        if not skip:
            scenarios.append({'name': name, 'filename': filename,
                'description': description})
    if scenarios:
        write_json(filename=outputdir/f'{prefix}_sources.json', d=scenarios)
    return scenarios, multipliers


def extract_custom_tla(wb, outputdir):
    """Extract custom TLA from an Excel file.
       Arguments:
         wb: Excel workbook as returned by openpyxl.
         outputdir: directory where output files are written
    """
    tla_tab = wb['TLA Data']
    title_cell = xls(tla_tab, 'A642')
    assert title_cell == 'Customized TLA Data', 'Title Cell: ' + title_cell
    assert xli(tla_tab, 'B645') == 2012

    df = pd.read_excel(wb, engine='openpyxl', sheet_name='TLA Data',
                       header=644, index_col=0, usecols="B:L", nrows=49)
    df.index.name = 'Year'
    df.index.astype(int)
    df = df.dropna(how='all', axis=1).dropna(how='all', axis=0)
    if df.empty:
        raise ValueError('Custom TLA is selected but there is no Custom TLA Data')
    else:
        df.to_csv(os.path.join(outputdir, 'custom_tla_data.csv'), index=True, header=True, encoding='utf-8')


def extract_vmas(f, wb, outputdir):
    """Extract VMAs from an Excel file.
       Arguments:
         f: output __init__.py file
         wb: Excel workbook as returned by openpyxl.
         outputdir: directory where output files are written
    """
    vma_dir_path = os.path.join(outputdir, 'vma_data')
    if not os.path.exists(vma_dir_path):
        os.mkdir(vma_dir_path)
    vma_r = VMAReader(wb)
    vmas = vma_r.read_xls(csv_path=vma_dir_path)
    vma_name_to_dict = {}
    for _, row in vmas.iterrows():
        vma_name = row['Title on xls']
        filename = row['Filename']

        if is_elecgen and 'CONVENTIONAL' in vma_name:
            continue    # don't list shared VMAs
        if not isinstance(filename, str) or not filename: 
            continue    # don't list non-existent VMAs

        vma_out_dict = {}
        vma_out_dict["filename"] = filename
        vma_out_dict['use_weight'] = row['Use weight?']
        vma_out_dict["bound_correction"] = row['Bound correction?']
        vma_out_dict["description"] = row['Description']
        vma_name_to_dict[vma_name] = vma_out_dict
    write_json(filename=Path(vma_dir_path)/'vma_sources.json', d=vma_name_to_dict)
    if is_elecgen:
        f.write("VMAs = (vma.VMA.load_vma_directory(THISDIR/'vma_data/vma_sources.json') | \n")
        f.write("        vma.VMA.load_vma_directory(DATADIR/'energy/vma_data/vma_sources.json'))\n")
    else:
        f.write("VMAs = vma.VMA.load_vma_directory(THISDIR/'vma_data/vma_sources.json')\n")
    f.write("\n")


def link_vma(tab, row, col):
    """
    Certain AdvancedControls inputs are linked to the mean, high or low value of their
    corresponding VMA tables. In the Excel ScenarioRecord, the cell value will look like:
    'Val:(328.415857769938) Formula:=C80'
    We can infer the chosen statistic from the cell reference. If there is no forumla we
    return the cell value as a float with no reference to the VMA result.
    Args:
      tab: the Sheet object to use
      row: numeric row number to look at
      col: numeric column number to look at

    Returns:
      'mean', 'high' or 'low' or raw value if no formula in cell
    """

    cell_value = xls(tab, row, col)
    if cell_value == '':
        return 0.0

    if 'Formula:=' not in cell_value:  # No formula present
        return xln(tab, row, col)
    else: # formula is present
        float_val = convert_sr_float(tab, row, col)

        # detect the standard statistics by the row number they reference
        if True in [cell_value.endswith(x) for x in ['80', '95', '101', '116', '146', '161', '175', '189', '140']]:
            return {'value': float_val, 'statistic': 'mean'}
        elif True in [cell_value.endswith(x) for x in ['81', '96', '102', '117', '147', '162', '176', '190', '141']]:
            return {'value': float_val, 'statistic': 'high'}
        elif True in [cell_value.endswith(x) for x in ['82', '97', '103', '118', '148', '163', '177', '191', '142']]:
            return {'value': float_val, 'statistic': 'low'}
        
        else: # not a standard formula, pass along the entire expression
            formula = cell_value.split(':=')[1]
            # Denise 7/21 turned off this warning because we get it too often.
            # Do a hack-y accumulator instead.
            #warnings.warn(f'formula "{formula}" in {col}:{str(row)} not recognised - using value')
            warn_counts['unknown_formula'] = warn_counts['unknown_formula'] + 1
            return {'value': float_val, 'xls cell formula': formula}


def write_units_rrs(f, wb):
    """Write out units for this solution."""
    sr_tab = wb['ScenarioRecord']
    f.write('units = {\n')
    for row in range(1, sr_tab.max_row):
        col_d = xls(sr_tab, row, co("D"))
        col_e = xls(sr_tab, row, co("E"))
        if col_d == 'Name of Scenario:' and 'TEMPLATE' not in col_e:
            f.write('    "implementation unit": "' + normalize_unit(sr_tab, row + 5, co("F")) + '",\n')
            f.write('    "functional unit": "' + normalize_unit(sr_tab, row + 7, co("F")) + '",\n')
            f.write('    "first cost": "' + normalize_unit(sr_tab, row + 16, co("F")) + '",\n')
            f.write('    "operating cost": "' + normalize_unit(sr_tab, row + 17, co("F")) + '",\n')
            break
    f.write('}\n\n')


def write_units_land(f, wb):
    """Write out units for this solution."""
    sr_tab = wb['ScenarioRecord']
    f.write('units = {\n')
    for row in range(1, sr_tab.max_row):
        col_d = xls(sr_tab, row, co("D"))
        col_e = xls(sr_tab, row, co("E"))
        if col_d == 'Name of Scenario:' and 'TEMPLATE' not in col_e:
            f.write('    "implementation unit": None,\n')
            f.write('    "functional unit": "' + normalize_unit(sr_tab, row + 5, co("F")) + '",\n')
            f.write('    "first cost": "' + normalize_unit(sr_tab, row + 12, co("F")) + '",\n')
            f.write('    "operating cost": "' + normalize_unit(sr_tab, row + 13, co("F")) + '",\n')
            break
    f.write('}\n\n')


def find_RRS_solution_category(wb):
    """Find the solution category (REPLACEMENT or REDUCTION)."""
    ac_tab = wb['Advanced Controls']
    possible = [('A157', 'A159'), ('A142', 'A144')]
    for labelcell, valuecell in possible:
        label = xls(ac_tab, labelcell)
        if 'Is this primarily a' in label:
            return xls(ac_tab, valuecell)
    return None


def _scenario_creation_date_from_str(s):
  date_format = "%Y-%m-%d %H:%M:%S"
  return datetime.datetime.strptime(s, date_format)

    
def _scenarios_from_ac_dir(ac_path):
    """Returns all scenarios in ac_path, and earliest creation date."""
    names = []
    creation_dates = []
    for jsonfile in ac_path.glob('*.json'):
        d = json.loads( jsonfile.read_text(encoding='utf-8') )
        if 'name' not in d:
            # Not in expected ac format.
            continue
        names.append(d['name'])

        if 'creation_date' not in d:
            # Not in expected ac format.
            continue
        creation_dates.append(_scenario_creation_date_from_str(
            d['creation_date']))

    if names:
        return names, min(creation_dates)
    else:
        return [], None



def json_dumps_default(obj):
    """Default function for json.dumps."""
    if isinstance(obj, np.integer):
        return int(obj)
    elif isinstance(obj, np.floating):
        return float(obj)
    elif isinstance(obj, np.ndarray):
        return obj.tolist()
    elif isinstance(obj, pd.DataFrame):
        return [[obj.index.name, *obj.columns.tolist()]] + obj.reset_index().values.tolist()
    elif isinstance(obj, pd.Series):
        return [[obj.index.name, obj.name]] + obj.reset_index().values.tolist()
    elif isinstance(obj, ac.SOLUTION_CATEGORY):
        return ac.solution_category_to_string(obj)
    else:
        raise TypeError('Unable to JSON encode: ' + repr(obj))

def write_json(filename, d):
    """Write out the given dict to the given Path."""
    with filename.open(mode='w', encoding='utf-8') as f:
        json.dump(obj=d, fp=f, indent=4, default=json_dumps_default)



warn_counts = {
    'unknown_formula': 0
}

def output_solution_python_file(outputdir, xl_filename):
    """Excel file -> generated Python code + many data files.

       Arguments:
         outputdir: directory to put output in.
         xl_filename: an Excel file to open, can be xls/xlsm/etc.
           Note that we cannot run Macros from xlsm files, only read values.
    """
    warn_counts['unknown_formula'] = 0 # reset counter
    global is_rrs
    global is_land
    global is_elecgen
 
    # We may get arguments as strings or PATH objects; make them strings here.
    outputdir = str(outputdir)
    xl_filename = str(xl_filename)

    if not os.path.exists(outputdir):
        os.mkdir(outputdir)
    py_filename = os.path.join(outputdir, '__init__.py')
    if os.path.exists(py_filename):
        py_filename = os.path.join(outputdir, '__init__UPDATED.py')
        print(f'Generating new code at {py_filename} - please merge by '
              'hand with __init__.py')

    wb = openpyxl.load_workbook(filename=xl_filename,data_only=True,keep_links=False)
    ac_tab = wb['Advanced Controls']

    if ('BIOSEQ' in xl_filename or 'PDLAND' in xl_filename or 'L-Use' in xl_filename or
            'AEZ Data' in wb.sheetnames):
        is_rrs = False
        is_land = True
        is_elecgen = False
    elif 'RRS' in xl_filename or 'TAM Data' in wb.sheetnames:
        is_rrs = True
        is_land = False
        is_elecgen = ('ElectricityGenerationSolution' in wb['Advanced Controls']['B1'].value)
    else:
        raise ValueError('Cannot determine solution category')
    has_tam = is_rrs

    # We take the adoption base year to be ref base year, which is also that HT expects.
    adoption_base_year = wb['Helper Tables']['B21'].value

    f = open(py_filename, 'w', encoding='utf-8')

    solution_name = xls(ac_tab, 'C40')
    f.write('#' + str(solution_name) + ' solution model.\n')
    f.write('#   Originally exported from: ' + os.path.basename(xl_filename) + '\n')
    f.write('\n')
    f.write('from pathlib import Path\n')
    f.write('import numpy as np\n')
    f.write('import pandas as pd\n')
    f.write('\n')
    f.write('from model import adoptiondata\n')
    f.write('from model import advanced_controls as ac\n')
    if is_land:
        f.write('from model import aez\n')
    f.write('from model import ch4calcs\n')
    f.write('from model import co2calcs\n')
    f.write('from model import customadoption\n')
    f.write('from model import dd\n')
    f.write('from model import emissionsfactors\n')
    f.write('from model import firstcost\n')
    f.write('from model import helpertables\n')
    f.write('from model import operatingcost\n')
    f.write('from model import s_curve\n')
    f.write('from model import scenario\n')
    f.write('from model import unitadoption\n')
    f.write('from model import vma\n')

    if has_tam:
        f.write('from model import tam\n')
    elif is_land:
        f.write('from model import tla\n')

    if is_rrs:
        f.write('from solution import rrs\n\n')
        sc = ac.string_to_solution_category(find_RRS_solution_category(wb=wb))
        solution_category = f'ac.SOLUTION_CATEGORY.{ac.solution_category_to_string(sc).upper()}'
        scenarios = get_rrs_scenarios(wb=wb, solution_category=sc)
    elif is_land:
        f.write('from model import conversions\n\n')
        sc = ac.string_to_solution_category('LAND')
        solution_category = 'ac.SOLUTION_CATEGORY.LAND'
        scenarios = get_land_scenarios(wb=wb, solution_category=sc)
    else:
        scenarios = {}

    f.write("DATADIR = Path(__file__).parents[2]/'data'\n")
    f.write("THISDIR = Path(__file__).parent\n")
    extract_vmas(f=f, wb=wb, outputdir=outputdir)
    if is_rrs:
        write_units_rrs(f=f, wb=wb)
    if is_land:
        write_units_land(f=f, wb=wb)
    f.write(f"name = '{solution_name}'\n")
    f.write(f"solution_category = {solution_category}\n")
    f.write("\n")

    is_protect = False
    has_harvest = False
    use_custom_tla = False
    for s in scenarios.values():
        if s.get('use_custom_tla', ''):
            if not 'custom_tla_fixed_value' in s:
                extract_custom_tla(wb, outputdir=outputdir)
            use_custom_tla = True
        if 'delay_protection_1yr' in s.keys():
            is_protect = True
        if 'carbon_not_emitted_after_harvesting' in s.keys():
            has_harvest = True

    p = Path(f'{outputdir}/ac')

    p.mkdir(parents=False, exist_ok=True)
    prev_scenarios, min_creation_date = _scenarios_from_ac_dir(p)
    for name, s in scenarios.items():
        if min_creation_date:
            creation_date = _scenario_creation_date_from_str(s['creation_date'])
            if creation_date < min_creation_date:
                print(f'Skipping scenario {name}, earlier than existing '
                      'scenarios.')
                continue
        fname = p.joinpath(re.sub(r"['\"\n()\\/\.]", "", name).replace(' ', '_').strip() + '.json')
        write_json(filename=fname, d=s)
    f.write("scenarios = ac.load_scenarios_from_json(directory=THISDIR/'ac', vmas=VMAs)\n")
    f.write("\n")

    f.write('# These are the "default" scenarios to use for each of the drawdown categories.\n')
    f.write('# They should be set to the most recent "official" set"\n')
    f.write('PDS1 = "NOT SET"\n')
    f.write('PDS2 = "NOT SET"\n')
    f.write('PDS3 = "NOT SET"\n\n')

    f.write(f"class Scenario(scenario.{'Land' if is_land else 'RRS'}Scenario):\n")
    f.write( "    name = name\n")
    f.write( "    units = units\n")
    f.write( "    vmas = VMAs\n")
    f.write( "    solution_category = solution_category\n")
    f.write( "    module_name = THISDIR.stem\n")
    f.write(f"    base_year = {adoption_base_year}\n")
    f.write( "\n")
    f.write( "    def __init__(self, scen=None):\n")
    f.write( "        # AC\n")
    f.write( "        self.initialize_ac(scen, scenarios, PDS2)\n")
    f.write( "\n")
    if has_tam:
        f.write("        # TAM\n")
        write_tam(f=f, wb=wb, outputdir=outputdir)
    elif is_land:
        f.write("        # TLA\n")
        write_aez(f=f, wb=wb, use_custom_tla=use_custom_tla)

    f.write("        # ADOPTION\n")

    # If needed, write out the data files, and emit the code to load the data from them
    write_ad(f=f, wb=wb, outputdir=outputdir)
    write_ca(case='PDS', f=f, wb=wb, outputdir=outputdir)
    write_ca(case='REF', f=f, wb=wb, outputdir=outputdir)

    f.write("        (ref_adoption_data_per_region,\n")
    f.write("         pds_adoption_data_per_region,\n")
    f.write("         pds_adoption_trend_per_region,\n")
    f.write("         pds_adoption_is_single_source) = self.initialize_adoption_bases()\n")
    f.write("\n")

    write_ht(f=f, wb=wb)

    f.write("        # DERIVED VALUES\n")
    f.write("\n")
    write_ef(f=f, wb=wb)
    write_ua(f=f, wb=wb)
    write_fc(f=f, wb=wb)
    write_oc(f=f, wb=wb)

    write_c2_c4(f=f, is_protect=is_protect, has_harvest=has_harvest)

    if is_rrs:
        f.write("        self.r2s = rrs.RRS(total_energy_demand=ref_tam_per_region.loc[2014, 'World'],\n")
        f.write("            soln_avg_annual_use=self.ac.soln_avg_annual_use,\n")
        f.write("            conv_avg_annual_use=self.ac.conv_avg_annual_use)\n")
        f.write("\n")

    f.close()
    if warn_counts['unknown_formula'] > 0:
        warnings.warn(f"Extraction encountered {warn_counts['unknown_formula']} unknown formulas in values on ScenarioRecord tab")



if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Create python Drawdown solution from Excel version.')
    parser.add_argument('excelfile', help='Excel filename to process.')
    parser.add_argument('--outputdir', default=None, help='Directory to write generated code and files to.  Defaults to the same directory as excelfile.')
    args = parser.parse_args(sys.argv[1:])

    excelfile = Path(args.excelfile).resolve()
    outputdir = Path(args.outputdir) if args.outputdir else excelfile.parent

    output_solution_python_file(outputdir=outputdir, xl_filename=excelfile)
