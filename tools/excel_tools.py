""" Utilities for accessing spreadsheets for use with / comparison to solutions """
import pandas as pd
import numpy as np
import openpyxl as oxl
import pytest
import zipfile
from pathlib import Path

# Historical note: the *_extract.py files do not use the utilities in this file.
# That is becaue they are older code, and it hasn't been worth refactoring them.

def get_from_excel(path, sheet, excel_range, column_names = None):
    """Return the requested data from the spreadsheet as a pandas Dataframe.
    This is convenient shorthand for a common case; use `pd.read_excel` and `openpyxl` directly for more varied situations.
    
    path: a string path to open, or a file-like object to read from
    sheet: the name or index of the sheet (tab) to read; if empty or None, the first sheet is assumed.
    excel_range: the range to read, as a string in standard xls form (e.g "A2:B24")
    column_names: a list of column names to use for the result.  If empty or None, the first row of the range
        is assumed to contain the column names (and that row is not included in the data rows)

    Note: if the excel range contains an index column, you can assign that after the fact like this::

        mydat = get_from_excel(...)
        mydat.index = mydat['column name']
    """
    (firstcol, firstrow, lastcol, lastrow) = oxl.utils.cell.range_boundaries(excel_range)
    nrows=lastrow-firstrow+1
    ncols=lastcol-firstcol+1
    
    if column_names:
        if len(column_names) != ncols:
            raise ValueError(f"Number of column names ({len(column_names)}) does not match number of columns to read ({ncols})")
        header = None
        skiprows = firstrow-1
        nrows = lastrow-firstrow+1
    else:
        header = firstrow-1
        skiprows = 0
        nrows = lastrow-firstrow
    
    wb = None
    try:
        wb = oxl.load_workbook(path, read_only=True, data_only=True, keep_links=False)
        return pd.read_excel(wb, sheet, 
            header=header, names=column_names, usecols=list(range(firstcol-1,lastcol)), 
            skiprows=skiprows, nrows=nrows, engine="openpyxl" )
    
    finally:
        if wb:
            wb.close()

# Project Drawdown column names can be long and cumbersome to type.  So here are a couple of functions to make handling them a bit easier

def rename_column(df:pd.DataFrame, i, new_name):
    """Rename the `i`'th column in `df` to `new_name`.
    Returns a _new_ DataFrame.  Reassign to keep the modified df::
    
        mydf = rename_column(mydf, 2, 'nice name')
    """
    return df.rename( columns={df.columns[i]: new_name} )


def rename_all_columns(df:pd.DataFrame, name_list):
    """Rename all the columns of `df`.  If any name in the list is None or NaN, that column
    is omitted from the result.
    Returns a _new_ DataFrame.  Reassign to keep the modified df::
    
        mydf = rename_all_columns(mydf, ['year', None, 'A', 'B', 'C'])
    """
    if len(name_list) != len(df.columns):
        raise ValueError(f"Number of column names provided ({len(name_list)}) doesn't match number of columns in df ({len(df.columns)})")
    
    keepcols = [ x for x in name_list if not pd.isna(x) ]
    mapcols = dict( zip( df.columns, name_list ) )
    mappeddf = df.rename(columns=mapcols)
    return mappeddf[keepcols]


def solution_expected_results(solution_name, scenario_name, sheet_name):
    """Return one sheet of the test results stored with the solution, for a specific
    scenario. Note this returns the entire sheet, not any specific table within a sheet.
    """
    zipfilename = Path('solution',solution_name,'testdata','expected.zip')   
    with zipfile.ZipFile(file=zipfilename) as zipf:
        with zipf.open(f'{scenario_name}/{sheet_name}') as fd:
            return pd.read_csv(fd, header=None)


def excel_range_from_df(df, excel_range, to_numeric=True):
    """Return a subset of a DataFrame expressed in Excel notation.
    Use this method when df holds an entire sheet of a spreadsheet.
    Usually you will _not_ want to include the header (column name) row in the df.
    """
    (firstcol, firstrow, lastcol, lastrow) = oxl.utils.cell.range_boundaries(excel_range)
    result = df.iloc[ firstrow-1:lastrow, firstcol-1:lastcol ]
    if to_numeric:
        result = result.apply(pd.to_numeric,errors='ignore')
    return result


def range_shift(excel_range, 
        row_shift=0, column_shift=0, 
        width_shift=0, height_shift=0,
        width_set=None, height_set=None):
    """Return an excel range with the specified modifications."""
    
    (firstcol, firstrow, lastcol, lastrow) = oxl.utils.cell.range_boundaries(excel_range)

    firstrow = firstrow + row_shift
    firstcol = firstcol + column_shift

    lastrow = firstrow + height_set - 1 if height_set else lastrow + row_shift + height_shift
    lastcol = firstcol + width_set - 1 if width_set else lastcol + column_shift + width_shift

    firstcname = oxl.utils.cell.get_column_letter(firstcol)
    lastcname = oxl.utils.cell.get_column_letter(lastcol)

    return f'{firstcname}{firstrow}:{lastcname}{lastrow}'
 

def approx_compare(val, expt, all_zero=True, thresh=None):
    """Return True if val is equal 'or very close to' expected value expt.
    Values must be numeric or strings.
    If all_zero is True (the default), 0, NaN, None and the empty string are all treated as equal.
    If thresh is provided, it overrides the default threshold to compare two floating point numbers.    
    """
    # This implementation is derived from the old test_excel_integration.compare_dataframes,
    # but it is not 100% identical.  The differences might matter later when we get deeper into
    # testing.

    pseudo_zero = lambda x : x == 0 or x == '' or x is None or x == 'NaN' or (isinstance(x,float) and np.isnan(x)) or x == pytest.approx(0.0)

    if isinstance(val, str) and isinstance(expt, str):
        return (val == expt)
    
    if all_zero and pseudo_zero(val):
        return pseudo_zero(expt)

    return (val == pytest.approx(expt) if thresh is None else
            val == pytest.approx(expt, abs=thresh, rel=1e-6))


def df_differ(val, expt, mask=None, all_zero=True, thresh=None):
    """Compare Dataframes val and expt using approximate comparison (see
    `excel_tools.approx_compare`).  The dataframes are compared by postition (index and column
    labels are ignored). If mask is provided, skip comparison for any mask cell that evals to
    True.  Parameters all_zero and thresh are passed directly to `approx_compare`.

    If the dataframes do differ, return a list of tuples `(row, col, val.value, expt.value)`
    for each cell that differs.  Return False if they do not differ.
    """
    if val.shape != expt.shape:
        raise ValueError(f'Dataframes differ in shape ({val.shape} vs {expt.shape})')

    result = []   
    (nrows,ncols) = val.shape
    for r in range(nrows):
        for c in range(ncols):
            if mask is not None and mask.iloc[r,c]:
                continue

            if not approx_compare(val.iloc[r,c], expt.iloc[r,c], all_zero=all_zero, thresh=thresh):
                result.append( (r, c, val.iloc[r,c], expt.iloc[r,c]) )
    
    return result if len(result) else False
