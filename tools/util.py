import re
import hashlib
import openpyxl
import pandas as pd
from numpy import nan
from pathlib import Path

##### Excel index manipulation

def co(colref):
    """Convert a column reference like "D" or "AA" to an index in 1-based notation.
    If colref is already an integer, passes it on unchanged."""
    return colref if isinstance(colref, int) else openpyxl.utils.cell.column_index_from_string(colref)

def cell_to_indices(cellref):
    """Convert an Excel reference like "C33" to (row, col) in 1-based notation (suitable for openpyxl)"""
    return openpyxl.utils.cell.coordinate_to_tuple(cellref)

def cell_to_offsets(cellref):
    """Convert an Excel reference like "C33" to (row, col) in 0-based notation (suitable for pd.read_csv or pd.read_excel"""
    ofs = cell_to_indices(cellref)
    return (ofs[0]-1, ofs[1]-1)

def df_excel_range(df, rangeref, to_numeric=True):
    """Return a subset of a DataFrame expressed in Excel notation.
    Use this method when df holds an entire sheet of a spreadsheet.
    Usually you will _not_ want to include the header (column name) row in the df.
    """
    (firstcol, firstrow, lastcol, lastrow) = openpyxl.utils.cell.range_boundaries(rangeref)
    result = df.iloc[ firstrow-1:lastrow, firstcol-1:lastcol ]
    if to_numeric:
        result = result.apply(pd.to_numeric,errors='ignore')
    return result

def find_in_column(ws, col, val, start_row=1, end_row=None):
    """Return the first row number (1-based) beginning with start_row where val is 
    found in the indicated column."""
    x = start_row
    end_row = end_row or ws.max_row
    while x <= end_row:
        if ws.cell(x,col).value == val:
            return x
        x = x+1
    return None

def find_in_row(ws, row, val, start_col=1, end_col=None):
    """Return the first column number (1-based) beginning with start_col where val is
    found in the indicated row."""
    x = start_col
    end_col = end_col or ws.max_column
    while x <= end_col:
        if ws.cell(row,x).value == val:
            return x
        x = x+1
    return None

def read_row(ws, row_num, start_col=1, end_col=None):
    """Return a list of values from ws for the given row (columns are inclusive)"""
    end_col = end_col or ws.max_column
    gen = ws.iter_rows(row_num, row_num, start_col, end_col, values_only=True)
    return list(next(gen))

def read_range(ws, start_row, start_col, end_row, end_col):
    """Return an array of arrays (row-oriented) of the data in the specified range of
    the worksheet (inclusive).  No data conversion is done."""
    return [ list(x) for x in ws.iter_rows(start_row, end_row, start_col, end_col, values_only=True) ]

##### Read Excel values, with type conversion
# The functions xls, xln and xli all take two forms of parameters
#   xls(tab, ref)        # ref in "A3" format
#   xls(tab, row, col)   # row and column in 1-based notation

def xls(tab, row, col=None):
    """Return a string value from tab(ref) or tab(row, col), where tab is a openpyxl sheet"""
    if col is None:
        (row, col) = cell_to_indices(row)
    val = tab.cell(row, col).value
    if val is None or tab.cell(row, col).data_type == 'e':
        return ''
    return str(val).strip()

def xln(tab, row, col=None, empty_is_nan=False):
    """Return the floating point number read from tab(ref) or tab(row, col), where tab is a openpyxl sheet.
    Returns NaN in case of error, and NaN or zero in case of empty, depending on `empty_is_nan` """
    if col is None:
        (row, col) = cell_to_indices(row)
    if tab.cell(row, col).data_type == 'e': # error
        return nan
    val = tab.cell(row, col).value
    if val is None or val == '':
        return nan if empty_is_nan else 0.0
    return float(val)

def xli(tab, row, col=None):
    """Return the integer read from tab(ref) or tab(row, col), where tab is a openpyxl sheet.
    Returns zero in case of empty or error."""
    if col is None:
        (row, col) = cell_to_indices(row)
    val = tab.cell(row, col).value
    if val is None or val == '' or tab.cell(row, col).data_type == 'e':
        return 0
    return int(val)

##### General Type conversions

def convert_bool(val, accept_empty=False):
    """Infer a boolean from common conventions in the spreadsheet."""
    if val is None:
        if accept_empty:
            return False
        raise ValueError('Cannot convert empty value to boolean')
    
    if isinstance(val, int):
        if val == 1:
            return True
        if val == 0:
            return False
        raise ValueError('Cannot convert arbitrary integer to boolean')    
    
    v = str(val).lower()
    if v in ['y','yes','1']:
        return True
    if v in ['n','no','0']:
        return False
    if accept_empty and v == '':
        return False
    raise ValueError('Unknown boolean format: ' + str(val))

def convert_float(val, return_nan=False):
    """Convert a value to floating point, with empty value ==> 0 or nan"""
    if val is None or val == '':
        return nan if return_nan else 0
    else:
        return float(val)

def normalize_region_name(name):
    """Correct common errors in region name"""
    fixes = {'Middle East & Africa': 'Middle East and Africa',
             'Asia (sans Japan)': 'Asia (Sans Japan)'}
    return fixes.get(name, name)

##### Conversion to filenames

def to_filename(title, prefix='', suffix='.csv', maxlen=35):
    """Convert a title (or any other string) into a string suitable for a filename."""
    # This currently allows any UTF-8 word characters through, which might cause problems on some
    # OS?, but as it hasn't been an issue yet, going to ignore it for now.
    # If we end up needing to do something, then look here:
    # https://stackoverflow.com/questions/295135/turn-a-string-into-a-valid-filename
    mlen = maxlen - len(prefix) - len(suffix)
    hashlen = 8  # how many characters of the hash to keep
    if mlen < hashlen+1:
        raise ValueError(f"Insufficient length {maxlen} in to_filename")

    # remove most punctuation
    filename = re.sub(r"[^\w\s\.\-]", '', title) 
    # replace sequences of whitespace and remaining punctuation with '_'
    filename = re.sub(r"[\s\.\-_]+", '_', filename)
    filename = filename.replace('Based_on_', 'bon_')  # specific to solution_xls names
    filename = filename.strip('_')
    
    if len(filename) > mlen:
        h = hashlib.sha256(title.encode('utf-8')).hexdigest()[-hashlen:]
        filename = filename[:mlen-hashlen-1] + '_' + h
    return prefix + filename + suffix

def to_unique_filename(title, existing_names, prefix='', suffix='.csv', maxlen=35):
    """Convert a title (or other language string) to a filename that is not already
    contained in the array of existing_names."""
    
    # look at the names only, not any directory info, and do comparisons in
    # lower case so that we don't end up with any names differing by case only.
    existing_names = [p.name if isinstance(p,Path) else p for p in existing_names]
    existing_names = [p.lower() for p in existing_names]
    
    p = Path(to_filename(title, prefix, suffix, maxlen))
    filename = p.name
    stem = p.stem
    cnt = 1
    while filename.lower() in existing_names:
        filename = stem + "_" + str(cnt) + suffix
        cnt += 1
    return filename
