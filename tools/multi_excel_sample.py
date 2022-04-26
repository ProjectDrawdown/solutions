""" Capture the same location within multiple spreadsheets for easy cross-model comparison.
Useful for determining if you are looking at a one-off issue, or for determining if some
old code is still needed.
Output is a spreadsheet containing one tab, with the requested section repeated for each
input file.  The section is repeated twice: first the data (which is useful to make sure
you've 'registered' the same location properly, and second with the formlas *as strings*.
"""
from pathlib import Path
import openpyxl
import sys
import warnings

# openpyxl conmplains every time it encounters any widget it doesn't understand, and we just don't care.
if not sys.warnoptions:
    warnings.simplefilter("ignore")


def copy_xls_region(cells, outsheet, startrow, startcol, data_mode, format_mode):
    """Copy the data from cells to outsheet, at the requested location, in the requested mode.
    data_mode: boolean for data: True copies the data in the cells, false copies the formulas
    format_mode: if True, copy cell formatting as well"""

    torow = startrow
    for cellrow in cells:
        torow = torow + 1
        tocol = startcol
        for cell in cellrow:
            outcell = outsheet.cell(torow, tocol)
            outcell.value = cell.value if data_mode else copy_formula(cell.value)
            if format_mode:
                copy_format(cell, outcell)
            tocol = tocol + 1


def copy_format(from_cell, to_cell):
    if from_cell.alignment: to_cell.alignment = from_cell.alignment
    if from_cell.border: to_cell.border = from_cell.border
    if from_cell.fill: to_cell.fill = from_cell.fill
    if from_cell.number_format: to_cell.number_format = from_cell.number_format


def copy_formula(value):
    """If the value is a formula, put a single quote in front of it to turn it into a string"""
    if (str(value).startswith('=')):
        return "'" + str(value)
    if value is None:
        return ""
    return str(value)


def find_cells(ws, textval, width, height):
    """Locate textval within ws and return a width-by-height selection of cells starting at that point"""
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == textval:
                return list(ws.iter_rows(cell.row, cell.row+height, cell.column, cell.column+width))
    raise ValueError("Text value not found")


def sample_regions(filelist, sheetname, region=None, copy_data=True, copy_formula=True, findtext=None, width=None, height=None):
    """For each file in the list, copy the specified sheet&region to a new Workbook.
    Returns the workbook.  
    Copies the data, formulas (as strings), or both."""

    resultwb = openpyxl.Workbook()
    resultsheet = resultwb[resultwb.sheetnames[0]] # get the first sheet
    resultsheet.title = sheetname + " " + (region.replace(":","-") if region else findtext[:10])

    # remove any "open excel" files from the filelist, and sort them.
    filelist = [ Path(x).resolve() for x in filelist ]
    filelist = sorted( [ x for x in filelist if not x.name.startswith('~')] )

    if region:
        (min_col, min_row, max_col, max_row) = openpyxl.utils.cell.range_boundaries(region)
        region_width = (max_col - min_col) + 1
        region_height = (max_row - min_row) + 1
    else:
        region_width = width
        region_height = height

    startcol = 2
    if copy_data:
        print("DATA")
        startrow = 2
        for file in filelist:
            print(file.name)
            # put file name in output sheet
            resultsheet.cell(startrow, 2).value = file.name
            src = openpyxl.load_workbook(file, read_only=True, data_only=True, keep_links = False)
            try:
                # get the requested cells, either by looking at the region, or finding the requested text
                src_sheet = src[sheetname]
                cells = src_sheet[region] if region else find_cells(src_sheet, findtext, width, height)
                copy_xls_region(cells, resultsheet, startrow+1, startcol, data_mode=True, format_mode=True)
            except Exception as e:
                resultsheet.cell(startrow+1, startcol).value = str(e)
            finally:
                startrow = startrow + region_height + 4
                src.close()
        
        startcol = startcol + region_width + 4
    
    if copy_formula:
        print("FORMULAS")
        startrow = 2
        for file in filelist:
            print(file.name)
            if startcol == 2:  # there was no data column, so output file name now.
                resultsheet.cell(startrow, 2).value = file.name
            src = openpyxl.load_workbook(file, read_only=True, keep_links = False)
            try:
                src_sheet = src[sheetname]
                cells = src_sheet[region] if region else find_cells(src_sheet, findtext, width, height)
                copy_xls_region(cells, resultsheet, startrow+1, startcol, data_mode=False, format_mode=False)
            except Exception as e:
                resultsheet.cell(startrow+1, startcol).value = str(e)            
            finally:
                startrow = startrow + region_height + 4
                src.close()

    return resultwb


if __name__ == "__main__":
    import argparse
    import glob
    import os

    parser = argparse.ArgumentParser(
        description='Sample the same region across multiple Excel workbooks')
    parser.add_argument('--sheet', help='Excel sheet to look at')
    parser.add_argument('--region', help='Region inside sheet to look at in Excel notation (e.g. B14:D15)')
    parser.add_argument('--out', required=False, default="sample.xlsx", help='Where to put the output Excel; defaults to sample.xlsx')
    parser.add_argument('files', nargs='+', help='Excel files to search')
    args = parser.parse_args()

    # Windows won't expand all glob'd paths, so do it ourselves.
    files = []
    for f in args.files:
        files.extend( glob.glob(f) )
    if len(files) == 0:
        print(f"**** No files matched {str(args.files)}")
        sys.exit()
    
    # Do checks on output file first, because otherwise you may spend a lot of time making the extract and
    # then lose it all at the end.
    outfile = Path(args.out)
    outdir = outfile.parent

    if outfile.suffix != ".xlsx":
        print("****  Setting extension of out file to .xlsx")
        outfile = Path(outfile).with_suffix('.xlsx')

    if not outdir.is_dir():
        print(f"****  directory {outdir.resolve()} does not exist")
        sys.exit(-1)
    if outfile.is_file():
        if not os.access(str(outfile), os.W_OK):
            print(f"**** cannot overwrite {outfile}; pick another file location!")
            sys.exit(-1)
    elif not os.access(str(outdir), os.W_OK):
        print(f"****  cannot write to directory {outdir}; pick another file location!")
        sys.exit(-1)

    wb = sample_regions(files, args.sheet, args.region)
    wb.save(outfile)
